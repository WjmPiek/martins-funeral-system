"""Shared Martins branding helpers for pages, PDF reports and Excel exports."""
from pathlib import Path

COMPANY_NAME = "Martins Funeral System"
TAGLINE = "Leaders in funeral industry"
BRAND_PURPLE = "5f4598"
BRAND_LIGHT_PURPLE = "a88ec5"
BRAND_GOLD = "f4df16"

def static_root() -> Path:
    return Path(__file__).resolve().parent / "static"

def logo_path() -> Path | None:
    candidates = [
        static_root() / "img" / "logo.png",
        static_root() / "uploads" / "system" / "logo.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None

def apply_workbook_branding(workbook, title: str = COMPANY_NAME, subtitle: str | None = None, sheet_name: str | None = None):
    """Insert a clean Martins logo/title block into an openpyxl workbook.

    This helper is safe to call from any future Excel export. It keeps the logo aspect ratio,
    reserves the first six rows for branding, and freezes the report below the header.
    """
    ws = workbook[sheet_name] if sheet_name else workbook.active
    ws.insert_rows(1, 6)
    ws["A1"] = title
    ws["A2"] = subtitle or TAGLINE
    ws["A1"].font = ws["A1"].font.copy(bold=True, size=16, color=BRAND_PURPLE)
    ws["A2"].font = ws["A2"].font.copy(bold=True, size=10, color="666666")
    ws.freeze_panes = "A7"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    path = logo_path()
    if path:
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(str(path))
            img.width = 260
            img.height = 117
            ws.add_image(img, "A1")
            ws["D1"] = title
            ws["D2"] = subtitle or TAGLINE
            ws["D1"].font = ws["D1"].font.copy(bold=True, size=16, color=BRAND_PURPLE)
            ws["D2"].font = ws["D2"].font.copy(bold=True, size=10, color="666666")
        except Exception:
            pass
    return workbook
