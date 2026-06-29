from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from openpyxl import load_workbook
from app import db
from app.models import PolicyProduct, PolicyChangeLog, ClientApplication, PolicyProductRule, Role, Permission, ClientApplication
from app.security import permission_required

policies_bp = Blueprint("policies", __name__, url_prefix="/policies")


def money(value):
    if value in (None, ""):
        return 0
    try:
        return float(str(value).replace("R", "").replace(",", "").strip())
    except Exception:
        return 0


def months_from_text(value):
    if value in (None, ""):
        return 0
    txt = str(value)
    digits = "".join(ch for ch in txt if ch.isdigit())
    return int(digits) if digits else 0


def age_min_max(value):
    if value in (None, ""):
        return (None, None)
    txt = str(value).replace("–", "-").replace("to", "-")
    nums = []
    current = ""
    for ch in txt:
        if ch.isdigit():
            current += ch
        elif current:
            nums.append(int(current))
            current = ""
    if current:
        nums.append(int(current))
    if len(nums) >= 2:
        return nums[0], nums[1]
    if len(nums) == 1:
        return nums[0], None
    return None, None


def is_yes(value):
    return str(value or "").strip().lower() in {"yes", "y", "true", "1", "x"}


@policies_bp.route("/")
@login_required
@permission_required("policies.view")
def list_products():
    products = PolicyProduct.query.order_by(PolicyProduct.product_name, PolicyProduct.plan_name).all()
    return render_template("policies/list.html", products=products)


@policies_bp.route("/import", methods=["GET", "POST"])
@login_required
@permission_required("policies.edit")
def import_policies():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Please choose the 2025 Policy Data Excel file.", "danger")
            return redirect(url_for("policies.import_policies"))

        wb = load_workbook(file, data_only=True)
        imported = 0
        updated = 0

        if "POLICIES" not in wb.sheetnames:
            flash("The workbook must contain a POLICIES sheet.", "danger")
            return redirect(url_for("policies.import_policies"))

        ws = wb["POLICIES"]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            product_name = str(data.get("ProductName") or "").strip()
            if not product_name:
                continue

            cover_amount = money(data.get("Cover Amount"))
            monthly_premium = money(data.get("Retail Premium"))
            waiting_months = months_from_text(data.get("Waiting Period"))
            min_age, max_age = age_min_max(data.get("Age Limit"))

            product = PolicyProduct.query.filter_by(product_name=product_name, plan_name=product_name).first()
            if not product:
                product = PolicyProduct(
                    product_name=product_name,
                    plan_name=product_name,
                    cover_amount=cover_amount,
                    monthly_premium=monthly_premium,
                    waiting_period_months=waiting_months,
                    min_age=min_age,
                    max_age=max_age,
                    active=True,
                )
                db.session.add(product)
                db.session.flush()
                imported += 1
            else:
                changes = {
                    "cover_amount": cover_amount,
                    "monthly_premium": monthly_premium,
                    "waiting_period_months": waiting_months,
                    "min_age": min_age,
                    "max_age": max_age,
                }
                for field, new_value in changes.items():
                    old_value = getattr(product, field)
                    if str(old_value) != str(new_value):
                        db.session.add(PolicyChangeLog(
                            product_id=product.id,
                            changed_by_id=current_user.id,
                            field_name=field,
                            old_value=str(old_value),
                            new_value=str(new_value),
                            reason="Policy Excel import",
                        ))
                        setattr(product, field, new_value)
                updated += 1

            rules = PolicyProductRule.query.filter_by(product_id=product.id).first()
            if not rules:
                rules = PolicyProductRule(product_id=product.id)
                db.session.add(rules)

            rules.qty_cover = str(data.get("QTY Cover") or "")
            rules.workers_policy_premium_per_1000 = money(data.get("Workers Policy Premium per R1000.00"))
            rules.suicide_waiting_period = str(data.get("Suicide Waiting Period") or "")
            rules.age_limit = str(data.get("Age Limit") or "")
            rules.joining_fee = money(data.get("Joining Fee"))
            rules.workers_policy_cover = money(data.get("Workers Policy Cover"))
            rules.main_member_cover = money(data.get("Main Member"))
            rules.spouse_cover = money(data.get("Spouse"))
            rules.extended_cover = money(data.get("Extended"))
            rules.member_0_5_product_only = money(data.get("0 - 5 Member + Product Only"))
            rules.member_6_70_product_only = money(data.get("6 - 70 Member + Product Only"))
            rules.stillborn_cover = money(data.get("Stillborn"))
            rules.family_0_11 = money(data.get("0 - 11 Family Plan Only"))
            rules.family_1_5 = money(data.get("1 - 5 Family Plan Only"))
            rules.family_6_13 = money(data.get("6 - 13 Family Plan Only"))
            rules.family_14_21 = money(data.get("14 - 21 Family Plan Only"))
            rules.reinstatement_rules = str(data.get("Reinstatement Rules") or "")
            rules.email_rule = str(data.get("Email") or "")
            rules.sms_whatsapp_rule = str(data.get("SMS/WhatsApp") or "")
            rules.require_draw_signature = is_yes(data.get("Signature method Draw signature on screen"))
            rules.require_typed_signature = is_yes(data.get("Signature method Type full name confirmation"))
            rules.require_otp = is_yes(data.get("Signature method OTP verification"))
            rules.document_storage = str(data.get("Document storage Server uploads folder") or "")
            rules.source_file = file.filename

        # Optional role update from USER ROLE sheet
        if "USER ROLE" in wb.sheetnames:
            ws_roles = wb["USER ROLE"]
            role_headers = [str(c.value).strip() if c.value is not None else "" for c in ws_roles[1]]
            for row in ws_roles.iter_rows(min_row=2, values_only=True):
                data = dict(zip(role_headers, row))
                role_name = str(data.get("User roles") or "").strip()
                if not role_name:
                    continue
                role = Role.query.filter(Role.name.ilike(role_name)).first()
                if not role:
                    role = Role(name=role_name, description=f"Imported from policy workbook")
                    db.session.add(role)
                    db.session.flush()

                base_permissions = []
                if is_yes(data.get("View")):
                    base_permissions += ["applications.view", "policies.view", "recovery.view"]
                if is_yes(data.get("Edit")):
                    base_permissions += ["applications.create", "policies.edit", "recovery.call", "recovery.import"]
                if is_yes(data.get("Email")):
                    base_permissions += ["applications.send_signing"]

                for code in set(base_permissions):
                    perm = Permission.query.filter_by(code=code).first()
                    if not perm:
                        perm = Permission(code=code, description=code)
                        db.session.add(perm)
                        db.session.flush()
                    if perm not in role.permissions:
                        role.permissions.append(perm)

        db.session.commit()
        flash(f"Policy import complete. New: {imported}. Updated/checked: {updated}.", "success")
        return redirect(url_for("policies.list_products"))

    return render_template("policies/import.html")


@policies_bp.route("/new", methods=["GET", "POST"])
@login_required
@permission_required("policies.edit")
def new_product():
    if request.method == "POST":
        p = PolicyProduct(
            product_name=request.form["product_name"], plan_name=request.form["plan_name"],
            cover_amount=request.form.get("cover_amount") or 0, monthly_premium=request.form.get("monthly_premium") or 0,
            waiting_period_months=request.form.get("waiting_period_months") or 0,
            min_age=request.form.get("min_age") or None, max_age=request.form.get("max_age") or None,
            active=bool(request.form.get("active"))
        )
        db.session.add(p); db.session.commit()
        flash("Policy product added", "success")
        return redirect(url_for("policies.list_products"))
    return render_template("policies/form.html", product=None)


@policies_bp.route("/<int:product_id>/edit", methods=["GET", "POST"])
@login_required
@permission_required("policies.edit")
def edit_product(product_id):
    p = PolicyProduct.query.get_or_404(product_id)
    if request.method == "POST":
        fields = ["product_name", "plan_name", "cover_amount", "monthly_premium", "waiting_period_months", "min_age", "max_age"]
        reason = request.form.get("reason", "Policy admin update")
        for f in fields:
            old = str(getattr(p, f))
            new = request.form.get(f) or None
            if old != str(new):
                db.session.add(PolicyChangeLog(product_id=p.id, changed_by_id=current_user.id, field_name=f, old_value=old, new_value=str(new), reason=reason))
                setattr(p, f, new)
        p.active = bool(request.form.get("active"))
        db.session.commit()
        flash("Policy updated and change log saved", "success")
        return redirect(url_for("policies.list_products"))
    return render_template("policies/form.html", product=p)


@policies_bp.route("/<int:product_id>/delete", methods=["POST"])
@login_required
@permission_required("policies.edit")
def delete_product(product_id):
    p = PolicyProduct.query.get_or_404(product_id)

    linked_apps = ClientApplication.query.filter_by(product_id=p.id).count()
    if linked_apps > 0:
        flash(f"Policy cannot be deleted because it is linked to {linked_apps} application(s). Edit it and mark it inactive instead.", "danger")
        return redirect(url_for("policies.list_products"))

    try:
        rules = PolicyProductRule.query.filter_by(product_id=p.id).first()
        if rules:
            db.session.delete(rules)

        # Delete change logs for this unused test/import policy
        PolicyChangeLog.query.filter_by(product_id=p.id).delete()

        db.session.delete(p)
        db.session.commit()
        flash("Policy product deleted.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Policy could not be deleted: {exc}", "danger")

    return redirect(url_for("policies.list_products"))
