from datetime import datetime, timedelta, date
from collections import defaultdict
from functools import wraps
import re
import secrets
import string
from difflib import SequenceMatcher
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort
from flask_login import login_required, current_user
from app.extensions import db
from app.models import User, Role, Permission, AuditLog, Franchise, RoyaltyScale, MonthlyFigure, user_franchises
from app.franchise_context import set_selected_franchise
from app.permissions import MODULES, ACTIONS, ROLE_TEMPLATES, ROLE_DEFAULTS, permission_code
from app.audit import log_action
from app.performance.service import auto_hide_inactive_franchises, inactive_franchise_candidates, reactivate_franchise_performance, has_recent_performance_data

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

PROTECTED_ADMIN_EMAIL = "wjm@martinsdirect.com"

ADMIN_SIDE_ROLE_NAMES = {"Admin", "Finance Manager", "Finance Assistant", "Regional Manager"}
FRANCHISE_SIDE_ROLE_NAMES = {"Franchise User", "Franchise Manager", "Franchise Employee", "Franchise Agent", "Read Only User"}
ADMIN_CREATABLE_ROLE_NAMES = ["Finance Manager", "Finance Assistant", "Regional Manager", "Franchise User"]
FRANCHISE_CREATABLE_ROLE_NAMES = ["Franchise Manager", "Franchise Employee", "Franchise Agent"]

ROLE_HELP_TEXT = {
    "Finance Manager": "Martins Funerals South Africa user. Sees the whole financial system and is not linked to one franchise.",
    "Finance Assistant": "Martins Funerals South Africa user. Finance support access and is not linked to one franchise.",
    "Regional Manager": "Martins Funerals South Africa user. Must be linked to the franchises/region they manage.",
    "Franchise User": "Franchise owner/user. Must be linked to the franchise(s) they own or operate.",
}
FINANCE_ADMIN_USERS = {
    "renette@martinsdirect.com": "Finance Manager",
    "lowhaan@martinsdirect.com": "Finance Assistant",
    "deon@martinsdirect.com": "Finance Assistant",
}


def ensure_user_hierarchy_roles():
    """Ensure the mother-company and franchise-level roles exist for the create-user screens."""
    descriptions = {
        "Finance Manager": "Martins Funerals South Africa finance manager",
        "Finance Assistant": "Martins Funerals South Africa finance assistant",
        "Regional Manager": "Martins regional manager linked to selected franchises",
        "Franchise User": "Franchise owner/user linked to selected franchise data",
        "Franchise Manager": "Manager created by a franchise user",
        "Franchise Employee": "Employee created by a franchise user",
        "Franchise Agent": "Agent created by a franchise user",
    }
    changed = False
    for role_name, description in descriptions.items():
        role = Role.query.filter_by(name=role_name).first()
        if not role:
            db.session.add(Role(name=role_name, description=description, is_system_role=True))
            changed = True
    if changed:
        db.session.flush()


def admin_creatable_roles():
    """Roles that Martins admin/finance users may create from the Admin Users page."""
    if is_current_user_admin():
        allowed = ADMIN_CREATABLE_ROLE_NAMES
    elif current_user.has_role("Finance Manager"):
        allowed = ["Finance Assistant", "Regional Manager", "Franchise User"]
    else:
        allowed = []
    return Role.query.filter(Role.name.in_(allowed)).order_by(Role.name).all()


def can_create_admin_user():
    # Admin must always be able to create Martins mother-company users.
    # Finance Manager may create allowed roles when Users Add permission is ticked.
    return is_current_user_admin() or (current_user.has_role("Finance Manager") and current_user.has_permission("users:add"))



def current_user_role_names():
    """Return role names robustly, including legacy/display role fields.

    Some older users show Admin in the UI but may not always have a populated
    role relationship in the current request. Treat the protected Martins admin
    account as Admin and read legacy single-role fields when present.
    """
    names = {role.name for role in getattr(current_user, "roles", []) or [] if getattr(role, "name", None)}

    for attr in ("role", "role_name", "user_role", "primary_role_name"):
        value = getattr(current_user, attr, None)
        if callable(value):
            try:
                value = value()
            except TypeError:
                value = None
        if value:
            names.add(str(value))

    email = (getattr(current_user, "email", "") or "").lower()
    full_name = (getattr(current_user, "full_name", "") or "").lower()
    if email == PROTECTED_ADMIN_EMAIL or full_name == "wjm piek":
        names.add("Admin")

    return names


def is_current_user_admin():
    return bool(current_user_role_names() & {"Admin", "Super Admin"})

def is_current_user_finance_import_user():
    return bool(current_user_role_names() & {"Finance Manager", "Finance Assistant"})


def can_view_imports_data():
    # Imports & Data is visible only to Admin and finance import users.
    # Admin sees all import tools. Finance Manager/Assistant see only Monthly Figures PDF.
    return is_current_user_admin() or is_current_user_finance_import_user()


def user_role_names(user):
    return {role.name for role in user.roles}


def is_admin_side_user(user):
    names = user_role_names(user)
    email = (user.email or "").lower()
    return email in FINANCE_ADMIN_USERS or bool(names & ADMIN_SIDE_ROLE_NAMES)


def is_franchise_side_user(user):
    names = user_role_names(user)
    return bool(names & FRANCHISE_SIDE_ROLE_NAMES) and not is_admin_side_user(user)


def is_role_admin_side(role_name):
    return role_name in ADMIN_SIDE_ROLE_NAMES


def role_requires_franchise_scope(role_name):
    return role_name in {"Regional Manager", "Franchise User"}


def normalise_user_scope_for_role(user, role_name, franchise_ids=None):
    """Keep admin-side users out of franchise hierarchy and scope franchise roles correctly."""
    user.parent_franchise_user_id = None

    if role_requires_franchise_scope(role_name):
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            return False, "Please link at least one active franchise for Regional Manager or Franchise User accounts."
        user.assigned_franchises = selected_franchises
        return True, ""

    # Finance/Admin-side users are Martins users. Finance Manager is linked to all active franchises.
    if is_role_admin_side(role_name):
        if role_name == "Finance Manager":
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
        return True, ""

    # Unknown roles are not allowed from the Admin create form.
    return False, "Please select a valid Admin-created role."


def tidy_finance_admin_users():
    # These named people are Martins Funerals South Africa/admin-side users, not franchise users.
    # Remove franchise links and accidental Franchise User/Manager roles, then keep their admin-side finance role.
    changed = 0
    for email, role_name in FINANCE_ADMIN_USERS.items():
        user = User.query.filter(db.func.lower(User.email) == email).first()
        if not user:
            continue
        wanted_role = Role.query.filter_by(name=role_name).first()
        cleaned_roles = [role for role in user.roles if role.name not in FRANCHISE_SIDE_ROLE_NAMES]
        if wanted_role and wanted_role not in cleaned_roles:
            cleaned_roles.append(wanted_role)
        if set(cleaned_roles) != set(user.roles):
            user.roles = cleaned_roles
            changed += 1
        if role_name == "Finance Manager":
            active_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
            if set(user.assigned_franchises) != set(active_franchises):
                user.assigned_franchises = active_franchises
                changed += 1
        elif user.assigned_franchises:
            user.assigned_franchises = []
            changed += 1
    return changed


def can_create_regional_manager():
    # Controlled by User Roles: grant Franchise Management Manage plus Users Add/Edit.
    return current_user.has_permission("franchise_management:manage") and (
        current_user.has_permission("users:add") or current_user.has_permission("users:edit")
    )


def can_assign_franchise_links():
    # Admin must always be able to link Regional Manager and Franchise User accounts.
    if is_current_user_admin() or current_user.has_role("Super Admin"):
        return True
    return current_user.has_permission("franchise_management:manage") and current_user.has_permission("users:edit")


def can_bulk_import_users():
    return current_user.has_permission("users:add") or current_user.has_permission("users:import")


def can_manage_old_franchises():
    names = current_user_role_names()
    return bool(names & {"Admin", "Super Admin", "Finance Manager"}) or current_user.has_permission("performance:manage_inactive")




def ordered_franchises_for_user(user):
    linked = list(getattr(user, "assigned_franchises", []) or [])
    if not linked:
        return []
    primary_id = db.session.execute(
        db.select(user_franchises.c.franchise_id)
        .where(user_franchises.c.user_id == user.id)
        .where(user_franchises.c.is_primary == True)
    ).scalar()
    linked_sorted = sorted(linked, key=lambda item: item.business_name or "")
    if primary_id:
        primary = [item for item in linked_sorted if item.id == primary_id]
        rest = [item for item in linked_sorted if item.id != primary_id]
        return primary + rest
    return linked_sorted

def active_linked_franchises_for_user(user):
    return [franchise for franchise in ordered_franchises_for_user(user) if getattr(franchise, "is_performance_active", True)]


def old_linked_franchises_for_user(user):
    return [franchise for franchise in ordered_franchises_for_user(user) if not getattr(franchise, "is_performance_active", True)]


def franchise_user_has_active_data(user):
    return bool(active_linked_franchises_for_user(user))


def franchise_user_has_recent_kpi_data(user, month=None, year=None):
    """A franchise user is listed only when at least one linked franchise has KPI data in the last 3 months."""
    now = datetime.utcnow()
    month = month or now.month
    year = year or now.year
    for franchise in active_linked_franchises_for_user(user):
        if has_recent_performance_data(franchise.id, month, year, 3):
            return True
    return False


def active_recent_franchise_owner_users(month=None, year=None):
    query_users = User.query.order_by(User.name, User.surname).all()
    return [
        user for user in query_users
        if user.has_role("Franchise User")
        and not getattr(user, "parent_franchise_user_id", None)
        and franchise_user_has_recent_kpi_data(user, month, year)
    ]


def is_protected_admin_user(user):
    return bool(user and (user.email or "").lower() == PROTECTED_ADMIN_EMAIL)



def can_change_user_roles():
    # Admin must always be able to correct user roles. Finance Assistant still
    # requires the Users Edit permission.
    names = user_role_names(current_user)
    if "Admin" in names or "Super Admin" in names:
        return True
    return "Finance Assistant" in names and current_user.has_permission("users:edit")


def clean_franchise_name(value):
    name = str(value or "").strip()
    if not name or name.upper() == "TOTAL":
        return ""
    return " ".join(name.split())


def slugify_email_part(value):
    value = value.lower().replace("&", "and")
    value = re.sub(r"\(f\)", "", value)
    value = re.sub(r"[^a-z0-9]+", ".", value)
    value = re.sub(r"\.+", ".", value).strip(".")
    return value or "franchise"


def temporary_password(length=14):
    alphabet = string.ascii_letters + string.digits + "!@#"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def get_or_create_role(role_name):
    role = Role.query.filter_by(name=role_name).first()
    if role:
        return role
    role = Role(name=role_name, description=f"Imported {role_name} role", is_system_role=True)
    db.session.add(role)
    db.session.flush()
    return role


def get_or_create_user(name, surname, email, role_name, franchises=None, password=None):
    email = email.strip().lower()
    user = User.query.filter_by(email=email).first()
    created = False
    if not user:
        user = User(name=name, surname=surname, email=email, is_active=True, is_active_account=True)
        user.set_password(password or temporary_password())
        db.session.add(user)
        db.session.flush()
        created = True
    else:
        user.name = user.name or name
        user.surname = user.surname or surname
        user.is_active = True
        user.is_active_account = True

    role = get_or_create_role(role_name)
    if role not in user.roles:
        user.roles.append(role)

    if franchises is not None:
        for franchise in franchises:
            if franchise not in user.assigned_franchises:
                user.assigned_franchises.append(franchise)
    return user, created


def permission_required(code):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Admin/Super Admin must never be blocked from Admin screens by
            # missing seeded permission rows or legacy user-role mismatches.
            if is_current_user_admin():
                return func(*args, **kwargs)
            if not current_user.has_permission(code):
                abort(403)
            return func(*args, **kwargs)
        return wrapper
    return decorator



def seed_permissions_and_roles():
    for module_index, module in enumerate(MODULES):
        for action_index, action in enumerate(ACTIONS):
            code = permission_code(module, action)
            permission = Permission.query.filter_by(code=code).first()
            if not permission:
                permission = Permission(
                    module=module,
                    action=action,
                    code=code,
                    label=f"{action.title()} {module}",
                    sort_order=(module_index * 100) + action_index,
                )
                db.session.add(permission)
    db.session.flush()

    all_permissions = Permission.query.all()
    permissions_by_code = {permission.code: permission for permission in all_permissions}

    for role_name, description in ROLE_TEMPLATES.items():
        role = Role.query.filter_by(name=role_name).first()
        if not role:
            role = Role(name=role_name, description=description, is_system_role=True)
            db.session.add(role)
            db.session.flush()
        defaults = ROLE_DEFAULTS.get(role_name, {})
        if defaults == "ALL":
            role.permissions = list(all_permissions)
        else:
            role.permissions = []
            for module, actions in defaults.items():
                for action in actions:
                    permission = permissions_by_code.get(permission_code(module, action))
                    if permission:
                        role.permissions.append(permission)
    db.session.commit()


@admin_bp.route("/seed")
@login_required
@permission_required("user_roles:manage")
def seed():
    seed_permissions_and_roles()
    flash("Default roles and permissions have been created/updated.", "success")
    return redirect(url_for("admin.roles"))


@admin_bp.route("/users")
@login_required
@permission_required("users:view")
def users():
    ensure_user_hierarchy_roles()
    db.session.commit()
    # Keep the franchise selector clean: branches with no KPI data in the last 3 months
    # are hidden automatically and shown in the Old Franchises tab until reactivated.
    now = datetime.utcnow()
    auto_hide_inactive_franchises(now.month, now.year, [franchise.id for franchise in Franchise.query.all()], current_user.id)
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    old_franchise_rows = inactive_franchise_candidates(now.month, now.year, [franchise.id for franchise in Franchise.query.order_by(Franchise.business_name).all()])
    old_franchises = [row for row in old_franchise_rows if not row["is_performance_active"]]
    selected_franchise_id = request.args.get("franchise_id", type=int)

    if selected_franchise_id:
        selected_franchise = Franchise.query.get_or_404(selected_franchise_id)
        if set_selected_franchise(selected_franchise.id, franchise_view_mode=True):
            flash(f"Opened {selected_franchise.business_name}. You can edit its details and royalty scale below.", "success")
            return redirect(url_for("franchise.details"))
        flash("You do not have access to that franchise.", "danger")
        return redirect(url_for("admin.users"))

    tidy_finance_admin_users()
    db.session.commit()

    all_users = User.query.order_by(User.name, User.surname).all()
    mother_company_users = [
        user for user in all_users
        if user.has_role("Admin")
        or user.has_role("Finance Manager")
        or user.has_role("Finance Assistant")
        or user.has_role("Regional Manager")
    ]
    franchise_owner_users = active_recent_franchise_owner_users(now.month, now.year)
    all_franchise_owner_users = [
        user for user in all_users
        if user.has_role("Franchise User") and not getattr(user, "parent_franchise_user_id", None)
    ]
    franchise_employee_users = [
        user for user in all_users
        if getattr(user, "parent_franchise_user_id", None)
        or user.has_role("Franchise Manager")
        or user.has_role("Franchise Employee")
        or user.has_role("Franchise Agent")
    ]
    admin_side_users = mother_company_users
    franchise_side_users = franchise_owner_users
    all_franchise_side_users = franchise_owner_users + franchise_employee_users
    old_franchise_users = [
        user for user in all_franchise_owner_users
        if ordered_franchises_for_user(user) and not franchise_user_has_recent_kpi_data(user, now.month, now.year)
    ]
    other_users = [user for user in all_users if user not in mother_company_users and user not in franchise_owner_users and user not in franchise_employee_users]
    linked_franchise_groups = []
    for user in franchise_owner_users:
        linked = ordered_franchises_for_user(user)
        if len(linked) > 1:
            linked_franchise_groups.append({"user": user, "main": linked[0], "franchises": linked})

    return render_template(
        "admin/users.html",
        users=all_users,
        admin_side_users=admin_side_users,
        franchise_side_users=franchise_side_users,
        mother_company_users=mother_company_users,
        franchise_owner_users=franchise_owner_users,
        franchise_employee_users=franchise_employee_users,
        other_users=other_users,
        old_franchise_users=old_franchise_users,
        linked_franchise_groups=linked_franchise_groups,
        roles=Role.query.order_by(Role.name).all(),
        franchises=franchises,
        old_franchises=old_franchises,
        selected_franchise=None,
        can_assign_franchise_links=can_assign_franchise_links(),
        can_change_user_roles=can_change_user_roles(),
        can_manage_old_franchises=can_manage_old_franchises(),
        admin_side_role_names=ADMIN_SIDE_ROLE_NAMES,
        franchise_side_role_names=FRANCHISE_SIDE_ROLE_NAMES,
        admin_creatable_roles=admin_creatable_roles(),
        can_create_admin_user=can_create_admin_user(),
        role_help_text=ROLE_HELP_TEXT,
    )


@admin_bp.route("/franchise-users")
@login_required
@permission_required("users:view")
def franchise_users():
    ensure_user_hierarchy_roles()
    now = datetime.utcnow()
    auto_hide_inactive_franchises(now.month, now.year, [franchise.id for franchise in Franchise.query.all()], current_user.id)
    db.session.commit()
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    franchise_users = active_recent_franchise_owner_users(now.month, now.year)
    return render_template(
        "admin/franchise_users.html",
        franchise_users=franchise_users,
        roles=Role.query.filter(Role.name.in_(["Franchise User"])).order_by(Role.name).all(),
        franchises=franchises,
        can_assign_franchise_links=can_assign_franchise_links(),
    )


@admin_bp.route("/users/create", methods=["GET", "POST"])
@login_required
def create_admin_user():
    if not can_create_admin_user():
        abort(403)

    if request.method == "GET":
        franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        return render_template(
            "admin/create_martins_user.html",
            admin_creatable_roles=admin_creatable_roles(),
            franchises=franchises,
        )

    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "").strip()
    role_id = request.form.get("role_id", type=int)
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]

    if not name or not surname or not email or not password or not role_id:
        flash("Name, surname, email, password and role are required.", "danger")
        return redirect(url_for("admin.users"))

    if User.query.filter(db.func.lower(User.email) == email).first():
        flash("A user with that email address already exists.", "danger")
        return redirect(url_for("admin.users"))

    role = Role.query.get_or_404(role_id)
    allowed_roles = {item.name for item in admin_creatable_roles()}
    if role.name not in allowed_roles:
        flash("You are not allowed to create that user role.", "danger")
        return redirect(url_for("admin.users"))

    user = User(
        name=name,
        surname=surname,
        email=email,
        is_active=True,
        is_active_account=True,
        parent_franchise_user_id=None,
        created_by_user_id=current_user.id,
    )
    user.set_password(password)
    user.roles.append(role)

    ok, message = normalise_user_scope_for_role(user, role.name, franchise_ids)
    if not ok:
        flash(message, "danger")
        return redirect(url_for("admin.users"))

    db.session.add(user)
    db.session.commit()
    log_action("Users", "Created admin-managed user", f"User: {email}; Role: {role.name}")
    flash(f"User {user.full_name} was created as {role.name}.", "success")
    if role.name == "Franchise User":
        return redirect(url_for("admin.franchise_users"))
    return redirect(url_for("admin.users"))


@admin_bp.route("/franchises/<int:franchise_id>/reactivate-performance", methods=["POST"])
@login_required
@permission_required("users:view")
def reactivate_old_franchise(franchise_id):
    if not can_manage_old_franchises():
        abort(403)
    franchise = reactivate_franchise_performance(franchise_id, current_user.id)
    if franchise:
        log_action("Franchise", "Reactivated old franchise", f"Franchise: {franchise.business_name}; ID: {franchise.id}")
        flash(f"{franchise.business_name} has been activated again and will be included in details, targets, graphs and calculations.", "success")
        return redirect(url_for("admin.users", franchise_id=franchise.id))
    flash("Franchise could not be found.", "danger")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/roles", methods=["POST"])
@login_required
def update_user_roles(user_id):
    if not can_change_user_roles():
        flash("Only Admin and Finance Assistant users with Users Edit permission may change user roles.", "danger")
        return redirect(url_for("admin.users"))

    user = User.query.get_or_404(user_id)

    if user.email and user.email.lower() == PROTECTED_ADMIN_EMAIL:
        admin_role = Role.query.filter_by(name="Admin").first()
        if admin_role and admin_role not in user.roles:
            user.roles.append(admin_role)
            db.session.commit()
        flash("Primary system administrator is protected. Roles cannot be changed.", "warning")
        return redirect(url_for("admin.users"))

    role_ids = [int(role_id) for role_id in request.form.getlist("role_ids")]
    selected_roles = Role.query.filter(Role.id.in_(role_ids)).all() if role_ids else []
    if any(role.name == "Regional Manager" for role in selected_roles) and not can_create_regional_manager():
        flash("Your role does not have permission to create or assign Regional Manager users.", "danger")
        return redirect(url_for("admin.users"))
    selected_role_names = {role.name for role in selected_roles}
    if "Admin" in selected_role_names and (user.email or "").lower() != PROTECTED_ADMIN_EMAIL:
        flash("The Admin role is locked to wjm@martinsdirect.com only.", "danger")
        return redirect(url_for("admin.users"))
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]

    # Mother-company finance/admin users must never sit under a franchise. Finance Manager is linked to all active franchise users/franchises.
    if selected_role_names & {"Admin", "Finance Manager", "Finance Assistant"}:
        user.parent_franchise_user_id = None
        if "Finance Manager" in selected_role_names:
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
    elif selected_role_names & {"Regional Manager", "Franchise User"}:
        user.parent_franchise_user_id = None
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            flash("Regional Manager and Franchise User accounts must be linked to at least one active franchise.", "danger")
            return redirect(url_for("admin.users"))
        user.assigned_franchises = selected_franchises
    else:
        # Admin > Users is only for Martins users and registered franchise owner/user accounts.
        # Franchise employees are managed separately under Admin > Employees and created by franchise owners.
        user.parent_franchise_user_id = None

    user.roles = selected_roles
    log_action("Users", "Updated user roles and scope", f"User: {user.full_name}")
    db.session.commit()
    flash(f"User and scope updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/update", methods=["POST"])
@login_required
@permission_required("users:edit")
def update_user(user_id):
    user = User.query.get_or_404(user_id)

    if is_protected_admin_user(user):
        flash("The primary Admin account is locked and cannot be edited.", "danger")
        return redirect(url_for("admin.users"))

    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    is_active = request.form.get("is_active") == "1"
    role_ids = [int(role_id) for role_id in request.form.getlist("role_ids") if str(role_id).isdigit()]
    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids") if str(item).isdigit()]

    if not name or not surname or not email:
        flash("Name, surname and email are required.", "danger")
        return redirect(url_for("admin.users"))

    duplicate = User.query.filter(db.func.lower(User.email) == email, User.id != user.id).first()
    if duplicate:
        flash("Another user already uses that email address.", "danger")
        return redirect(url_for("admin.users"))

    selected_roles = Role.query.filter(Role.id.in_(role_ids)).all() if role_ids else []
    if not selected_roles:
        flash("Please select at least one role.", "danger")
        return redirect(url_for("admin.users"))

    selected_role_names = {role.name for role in selected_roles}
    if "Admin" in selected_role_names and email != PROTECTED_ADMIN_EMAIL:
        flash("The Admin role is locked to wjm@martinsdirect.com only.", "danger")
        return redirect(url_for("admin.users"))
    if "Regional Manager" in selected_role_names and not can_create_regional_manager():
        flash("Your role does not have permission to assign Regional Manager users.", "danger")
        return redirect(url_for("admin.users"))

    user.name = name
    user.surname = surname
    user.email = email
    user.is_active = is_active
    user.is_active_account = is_active

    if selected_role_names & {"Admin", "Finance Manager", "Finance Assistant"}:
        user.parent_franchise_user_id = None
        if "Finance Manager" in selected_role_names:
            user.assigned_franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
        else:
            user.assigned_franchises = []
    elif selected_role_names & {"Regional Manager", "Franchise User"}:
        user.parent_franchise_user_id = None
        selected_franchises = Franchise.query.filter(
            Franchise.id.in_(franchise_ids or []),
            Franchise.is_performance_active == True,
        ).order_by(Franchise.business_name).all()
        if not selected_franchises:
            flash("Regional Manager and Franchise User accounts must be linked to at least one active franchise.", "danger")
            return redirect(url_for("admin.users"))
        user.assigned_franchises = selected_franchises
    else:
        user.parent_franchise_user_id = None

    user.roles = selected_roles
    log_action("Users", "Updated user details", f"User: {user.full_name}; Email: {user.email}")
    db.session.commit()
    flash(f"User updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@permission_required("users:delete")
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if is_protected_admin_user(user):
        flash("The primary Admin account is locked and cannot be deleted or deactivated.", "danger")
        return redirect(url_for("admin.users"))

    user.is_active = False
    log_action("Users", "Deactivated user", f"User: {user.full_name}")
    db.session.commit()
    flash(f"{user.full_name} has been deactivated. User data has been kept.", "success")
    return redirect(url_for("admin.users"))



@admin_bp.route("/users/cleanup-inactive", methods=["POST"])
@login_required
@permission_required("users:delete")
def cleanup_inactive_users():
    cutoff = datetime.utcnow() - timedelta(days=60)
    users = User.query.filter(User.email != PROTECTED_ADMIN_EMAIL).all()
    count = 0

    for user in users:
        last_seen = getattr(user, "last_login_at", None) or getattr(user, "created_at", None)
        if user.is_active and last_seen and last_seen < cutoff:
            user.is_active = False
            count += 1

    db.session.commit()
    log_action("Users", "Deactivated inactive users", f"Count: {count}")
    flash(f"{count} inactive user(s) deactivated. Their data was kept.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/import-franchise-users", methods=["GET", "POST"])
@login_required
def import_franchise_users():
    if not is_current_user_admin():
        abort(403)

    allowed_roles = ["Franchise User", "Franchise Manager"]
    if can_create_regional_manager():
        allowed_roles.append("Regional Manager")

    if request.method == "POST":
        target_role = request.form.get("target_role", "Franchise User").strip()
        if target_role not in allowed_roles:
            flash("You are not allowed to create that user role.", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the Excel file with franchise names in row 1.", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
            raw_names = [cell.value for cell in worksheet[1]]
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_franchise_users"))

        franchise_names = []
        seen = set()
        removed_totals = 0
        for value in raw_names:
            cleaned = clean_franchise_name(value)
            if not cleaned:
                if str(value or "").strip().upper() == "TOTAL":
                    removed_totals += 1
                continue
            key = cleaned.lower()
            if key not in seen:
                franchise_names.append(cleaned)
                seen.add(key)

        if not franchise_names:
            flash("No franchise names were found in row 1 after removing TOTAL columns.", "warning")
            return redirect(url_for("admin.import_franchise_users"))

        franchises = []
        franchises_created = 0
        users_created = 0
        users_updated = 0
        generated = []

        for franchise_name in franchise_names:
            franchise = Franchise.query.filter(db.func.lower(Franchise.business_name) == franchise_name.lower()).first()
            if not franchise:
                franchise = Franchise(business_name=franchise_name, franchise_code=slugify_email_part(franchise_name).upper()[:20])
                db.session.add(franchise)
                db.session.flush()
                franchises_created += 1
            franchises.append(franchise)

            email = f"{slugify_email_part(franchise_name)}@martinsdirect.com"
            password = temporary_password()
            user, created = get_or_create_user(franchise_name, "User", email, target_role, [], password)
            if created:
                users_created += 1
                generated.append((user.full_name, user.email, password, "Not linked - assign manually", target_role))
            else:
                users_updated += 1

        # Legacy behaviour removed: the import must not create Renette, Lowhaan or Deon as separate
        # Finance Manager / Finance Assistant users. Finance permissions now belong to the Admin role
        # or to roles explicitly assigned in User Roles.
        finance_created = 0
        finance_updated = 0
        db.session.commit()
        log_action(
            "Users",
            "Imported franchise users from Excel",
            f"Franchises: {len(franchises)}, new franchises: {franchises_created}, new users: {users_created}, updated users: {users_updated}, totals removed: {removed_totals}",
        )

        flash(
            f"Import complete. {len(franchises)} franchises processed, {franchises_created} new franchises created, "
            f"{users_created} franchise users created, {users_updated} users updated, {removed_totals} TOTAL columns removed. "
            "No separate finance users were created.",
            "success",
        )

        return render_template(
            "admin/import_franchise_users.html",
            allowed_roles=allowed_roles,
            selected_role=target_role,
            generated=generated,
            import_complete=True,
        )

    return render_template(
        "admin/import_franchise_users.html",
        allowed_roles=allowed_roles,
        selected_role="Franchise User",
        generated=[],
        import_complete=False,
    )



def normalize_franchise_key(value):
    """Create a forgiving key for matching spreadsheet branch names to Franchise records."""
    text = str(value or "").strip().lower()
    text = re.sub(r"\(f\)", "", text)
    text = text.replace("martin's", "martins")
    text = text.replace("martins funerals", "")
    text = text.replace("martins funeral", "")
    text = text.replace("martins begrafnisdienste", "")
    text = text.replace("begrafnisdienste", "")
    text = text.replace("funerals", "")
    text = re.sub(r"\bpty\b|\bltd\b|\blimited\b|\btas\b|\bt/a\b|\bck\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


AFRIKAANS_MONTHS = {
    "januarie": 1, "jan": 1,
    "februarie": 2, "feb": 2,
    "maart": 3, "mrt": 3,
    "april": 4, "apr": 4,
    "mei": 5,
    "junie": 6, "jun": 6,
    "julie": 7, "jul": 7,
    "augustus": 8, "aug": 8,
    "september": 9, "sep": 9,
    "oktober": 10, "okt": 10,
    "november": 11, "nov": 11,
    "desember": 12, "des": 12,
}


def parse_contract_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace("  ", " ")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = re.match(r"^(\d{1,2})\s+([A-Za-zÀ-ÿ]+)\s+(\d{4})$", text)
    if match:
        day = int(match.group(1))
        month = AFRIKAANS_MONTHS.get(match.group(2).lower())
        year = int(match.group(3))
        if month:
            return date(year, month, day)
    return None


def clean_excel_text(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def unique_join(values):
    seen = set()
    result = []
    for value in values:
        text = clean_excel_text(value)
        if not text:
            continue
        key = text.lower()
        if key not in seen:
            result.append(text)
            seen.add(key)
    return "; ".join(result)


def last_non_empty(values):
    chosen = ""
    for value in values:
        text = clean_excel_text(value)
        if text:
            chosen = text
    return chosen


def last_contract_date(values):
    chosen = None
    for value in values:
        parsed = parse_contract_date(value)
        if parsed:
            chosen = parsed
    return chosen


def newest_contract_date(values):
    """Return the newest valid date from a set of spreadsheet cells.

    Contract summary imports are a full refresh for matched franchises.
    When the Excel sheet is updated and uploaded again, the database must be
    overwritten from the new file instead of keeping older dates.
    """
    parsed_dates = []
    for value in values:
        parsed = parse_contract_date(value)
        if parsed:
            parsed_dates.append(parsed)
    return max(parsed_dates) if parsed_dates else None


def set_auto_gross_method_from_agreement(franchise):
    franchise.royalty_gross_method = "new" if (franchise.agreement_start_date and franchise.agreement_start_date.year >= 2018) else "old"


def sync_royalty_scales_from_contract_file(franchise, parsed_rows, raw_scale_lines, minimum):
    """Fully sync royalty-scale fields from the latest uploaded Contract Summary file.

    This intentionally replaces the database values every time the Excel file is
    uploaded. It avoids the old behaviour where blank/changed cells left old
    values behind and made the Franchise Details page look unchanged.
    """
    franchise.imported_royalty_scale_text = "\n".join(raw_scale_lines or [])
    franchise.imported_royalty_percentage = (parsed_rows[0].get("percentage") if parsed_rows else 0) or 0
    franchise.minimum_royalty_amount = minimum if minimum is not None else 0

    RoyaltyScale.query.filter_by(franchise_id=franchise.id).delete()
    db.session.flush()
    for index, parsed in enumerate(parsed_rows or [], start=1):
        db.session.add(RoyaltyScale(
            franchise_id=franchise.id,
            row_number=index,
            amount_from=parsed.get("amount_from") or 0,
            amount_to=parsed.get("amount_to") or 999999999,
            percentage=parsed.get("percentage") or 0,
        ))
    return len(parsed_rows or [])


def parse_money_token(value):
    if value is None:
        return None
    text = str(value)
    text = re.sub(r"[^0-9.,-]", "", text).replace(" ", "")
    if not text:
        return None
    if text.count(",") == 1 and text.count(".") == 0:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_royalty_scale_line(value):
    raw = clean_excel_text(value)
    if not raw:
        return None
    percent_match = re.search(r"(\d+(?:[.,]\d+)?)\s*%", raw)
    minimum_match = re.search(r"minimum.*?r\s*([0-9\s.,]+)", raw, re.I)
    if minimum_match:
        return {"raw": raw, "minimum": parse_money_token(minimum_match.group(1))}
    if not percent_match:
        return {"raw": raw}
    percentage = parse_money_token(percent_match.group(1)) or 0
    before_percent = raw[:percent_match.start()].strip(" -")
    money_values = [parse_money_token(item) for item in re.findall(r"R\s*[0-9][0-9\s.,]*", before_percent, flags=re.I)]
    money_values = [item for item in money_values if item is not None]
    if len(money_values) >= 2:
        amount_from, amount_to = money_values[0], money_values[1]
    elif len(money_values) == 1:
        if re.search(r"up to|tot en met", raw, re.I):
            amount_from, amount_to = 0, money_values[0]
        elif re.search(r"or more|and more|more|meer", raw, re.I):
            amount_from, amount_to = money_values[0], 999999999
        else:
            amount_from, amount_to = 0, money_values[0]
    else:
        amount_from, amount_to = 0, 999999999
    return {
        "raw": raw,
        "amount_from": amount_from,
        "amount_to": amount_to,
        "percentage": percentage,
        "label": before_percent,
    }


def find_franchise_by_name(name):
    key = normalize_franchise_key(name)
    if not key:
        return None
    franchises = Franchise.query.all()
    # exact normalized match first
    for franchise in franchises:
        if normalize_franchise_key(franchise.business_name) == key:
            return franchise
    # compare common shortened branch names
    for franchise in franchises:
        existing = normalize_franchise_key(franchise.business_name)
        if key and existing and (key == existing or key in existing or existing in key):
            return franchise
    return None




def split_grouped_franchise_names(value):
    """Split one grouped-franchise spreadsheet cell into ordered franchise names.

    The first name in the cell is the main franchise.  Remaining names are linked
    to that main franchise user for grouped royalty calculation.
    """
    raw = clean_excel_text(value)
    if not raw:
        return []
    parts = re.split(r"[,;\n]+", raw)
    result = []
    seen = set()
    for part in parts:
        name = clean_franchise_name(part)
        if not name:
            continue
        key = normalize_franchise_key(name)
        if not key or key in seen:
            continue
        result.append(name)
        seen.add(key)
    return result


def find_franchise_user_for_main_franchise(main_franchise):
    """Find or create the franchise-side user that owns a grouped royalty set.

    Important: a franchise may already be linked to another user's group.  For
    grouped-franchise imports the FIRST name in the spreadsheet row must become
    the main owner, so do not simply return any user linked to the franchise.
    Prefer the user's own generated email/name and only accept an existing
    assigned user if that franchise is already marked primary for that user.
    """
    franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
    email = f"{slugify_email_part(main_franchise.business_name)}@martinsdirect.com"

    user = User.query.filter(db.func.lower(User.email) == email.lower()).first()
    if user and not is_admin_side_user(user):
        role = get_or_create_role("Franchise User")
        if role not in user.roles:
            user.roles.append(role)
        return user, False

    normalized_main = normalize_franchise_key(main_franchise.business_name)
    for user in getattr(main_franchise, "assigned_users", []) or []:
        if not (user_role_names(user) & franchise_side_roles) or is_admin_side_user(user):
            continue
        primary_id = db.session.execute(
            db.select(user_franchises.c.franchise_id)
            .where(user_franchises.c.user_id == user.id)
            .where(user_franchises.c.is_primary == True)
        ).scalar()
        if primary_id == main_franchise.id:
            return user, False
        # Fallback for older data where is_primary was not set but the user's
        # name/email clearly belongs to the main franchise.
        user_key = normalize_franchise_key(f"{user.name} {user.surname}")
        if normalized_main and normalized_main in user_key:
            return user, False

    display_name = re.sub(r"\s*\(F\)\s*$", "", main_franchise.business_name or "Franchise", flags=re.I).strip() or "Franchise"
    user, created = get_or_create_user(display_name, "User", email, "Franchise User", franchises=[main_franchise])
    return user, created


def remove_group_links_from_other_franchise_users(user, franchise_ids):
    """Remove grouped-branch links from other franchise-side users before re-import.

    This prevents old incorrect groups (for example Dobsonville owning Soweto)
    from remaining after the spreadsheet says Soweto is the main franchise.
    Admin/Finance users are not touched.
    """
    if not franchise_ids:
        return
    franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
    franchise_side_user_ids = [
        row[0]
        for row in db.session.query(User.id)
        .join(User.roles)
        .filter(Role.name.in_(franchise_side_roles))
        .all()
    ]
    if not franchise_side_user_ids:
        return
    db.session.execute(
        user_franchises.delete()
        .where(user_franchises.c.franchise_id.in_(franchise_ids))
        .where(user_franchises.c.user_id != user.id)
        .where(user_franchises.c.user_id.in_(franchise_side_user_ids))
    )


def set_primary_franchise_link(user, main_franchise, linked_franchises):
    """Assign linked franchises and mark the main franchise as primary in user_franchises."""
    ordered = []
    seen = set()
    for franchise in [main_franchise] + list(linked_franchises or []):
        if franchise and franchise.id not in seen:
            ordered.append(franchise)
            seen.add(franchise.id)
    remove_group_links_from_other_franchise_users(user, list(seen))
    user.assigned_franchises = ordered
    db.session.flush()
    db.session.execute(user_franchises.update().where(user_franchises.c.user_id == user.id).values(is_primary=False))
    db.session.execute(
        user_franchises.update()
        .where(user_franchises.c.user_id == user.id)
        .where(user_franchises.c.franchise_id == main_franchise.id)
        .values(is_primary=True)
    )


def clean_contact_branch_name(value):
    """Clean contact-list branch/outlet names before matching to Franchise.business_name."""
    text = normalize_contact_value(value)
    text = text.replace("*", "").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def contact_candidate_names(branch, outlet):
    """Return forgiving franchise-name candidates from contact list columns B and C.

    Column B contains the main franchise/area and column C sometimes contains the
    actual outlet.  For outlet rows such as CAPE TOWN / PAROW, the existing
    Franchise record may be stored as either 'Parow', 'Cape Town Parow', or
    'Cape Town (Parow)', so all useful variants are tried.
    """
    branch = clean_contact_branch_name(branch)
    outlet = clean_contact_branch_name(outlet)
    candidates = []
    if branch and outlet:
        # Try the outlet first because the system franchise name is commonly the
        # outlet name, then fall back to combined names and the main branch.
        candidates.extend([
            outlet,
            f"{branch} {outlet}",
            f"{branch} - {outlet}",
            f"{branch} ({outlet})",
            branch,
        ])
    elif outlet:
        candidates.append(outlet)
    elif branch:
        candidates.append(branch)
    # Remove duplicates while preserving order.
    seen = set()
    result = []
    for item in candidates:
        key = normalize_franchise_key(item)
        if key and key not in seen:
            result.append(item)
            seen.add(key)
    return result


def find_franchise_by_candidates(candidates):
    """Find a franchise using exact, contains, and safe fuzzy matching.

    Contact-list branch names do not always match the saved franchise name exactly.
    Examples: '*CAPE TOWN' + 'PAROW' may be saved as 'Cape Town Parow',
    and some saved names include 'User', '(F)', 'Martin's Funerals', etc.
    """
    cleaned_candidates = []
    seen = set()
    for candidate in candidates:
        key = normalize_franchise_key(candidate)
        if key and key not in seen:
            cleaned_candidates.append((candidate, key))
            seen.add(key)

    franchises = Franchise.query.all()
    franchise_keys = [(franchise, normalize_franchise_key(franchise.business_name)) for franchise in franchises]

    # 1) Exact normalized match.
    for candidate, key in cleaned_candidates:
        for franchise, existing_key in franchise_keys:
            if key == existing_key:
                return franchise, candidate

    # 2) Containment match. Prefer longer keys to avoid broad matches.
    for candidate, key in sorted(cleaned_candidates, key=lambda item: len(item[1]), reverse=True):
        for franchise, existing_key in franchise_keys:
            if key and existing_key and (key in existing_key or existing_key in key):
                return franchise, candidate

    # 3) Safe fuzzy match for spelling/spacing variations.
    best = (None, "", 0.0)
    for candidate, key in cleaned_candidates:
        for franchise, existing_key in franchise_keys:
            if not key or not existing_key:
                continue
            ratio = SequenceMatcher(None, key, existing_key).ratio()
            if ratio > best[2]:
                best = (franchise, candidate, ratio)
    if best[0] is not None and best[2] >= 0.86:
        return best[0], f"{best[1]} (fuzzy {best[2]:.0%})"

    return None, ""




def clean_shareholder_name(value):
    """Return a clean shareholder name from column N.

    The contract sheet often stores shareholders as numbered values such as
    '1. Jan van Wyk'. Only the primary/first shareholder is imported into the
    single Franchisee Name/Surname fields.
    """
    text = clean_excel_text(value)
    if not text:
        return ""
    text = text.replace("\r", "\n").replace(";", "\n")
    candidates = []
    for part in re.split(r"\n+|(?=\b\d+\s*[.)]\s+)", text):
        part = clean_excel_text(part)
        if part:
            candidates.append(part)
    if not candidates:
        candidates = [text]
    # Prefer the explicitly numbered primary shareholder, otherwise use the first non-empty value.
    chosen = ""
    for candidate in candidates:
        if re.match(r"^1\s*[.)]\s+", candidate):
            chosen = candidate
            break
    if not chosen:
        chosen = candidates[0]
    chosen = re.sub(r"^\s*\d+\s*[.)]\s*", "", chosen).strip()
    return clean_excel_text(chosen)


def first_shareholder_name(values):
    for value in values:
        name = clean_shareholder_name(value)
        if name:
            return name
    return ""

def split_person_name(full_name):
    text = clean_excel_text(full_name)
    # The contact sheet sometimes includes a 24H number inside NAME & SURNAME.
    text = re.sub(r"24\s*h.*$", "", text, flags=re.I).strip()
    if not text:
        return "", ""
    parts = text.split()
    if len(parts) == 1:
        return parts[0].title(), ""
    return " ".join(parts[:-1]).title(), parts[-1].title()


def normalize_contact_value(value):
    text = clean_excel_text(value)
    text = text.replace(" ", " ")
    return " ".join(text.split())

def build_contract_summary_groups(worksheet):
    """Build current contract-summary groups from B2:B375.

    The contract workbook stores royalty-scale continuation rows with a blank
    franchise name in column B.  A franchise can also appear more than once when
    an old agreement and a newer agreement are both in the file.  In that case
    the import must keep the newest agreement block and replace the older data,
    otherwise the database will not reflect the latest Excel upload.
    """
    blocks = []
    current_block = None
    last_row = min(worksheet.max_row, 375)

    for row_number in range(2, last_row + 1):
        franchise_name = clean_excel_text(worksheet.cell(row_number, 2).value)
        if franchise_name:
            key = normalize_franchise_key(franchise_name)
            if not key or key in {"total", "totals", "data"}:
                current_block = None
                continue
            current_block = {"name": franchise_name, "key": key, "rows": []}
            blocks.append(current_block)

        if current_block:
            row_values = [worksheet.cell(row_number, c).value for c in (4, 5, 10, 11, 12, 13, 14, 18)]
            if franchise_name or any(clean_excel_text(v) for v in row_values):
                current_block["rows"].append(row_number)

    selected_by_key = {}
    selected_order = []
    for block in blocks:
        rows = block["rows"]
        block_start_date = newest_contract_date(worksheet.cell(r, 4).value for r in rows)
        existing = selected_by_key.get(block["key"])
        if not existing:
            block["start_date_for_selection"] = block_start_date
            selected_by_key[block["key"]] = block
            selected_order.append(block["key"])
            continue

        existing_date = existing.get("start_date_for_selection")
        should_replace = False
        if block_start_date and existing_date:
            should_replace = block_start_date >= existing_date
        elif block_start_date and not existing_date:
            should_replace = True

        if should_replace:
            block["start_date_for_selection"] = block_start_date
            selected_by_key[block["key"]] = block

    return [selected_by_key[key] for key in selected_order if key in selected_by_key]



@admin_bp.route("/imports")
@login_required
def imports_data():
    if not can_view_imports_data():
        abort(403)
    role_names = current_user_role_names()
    return render_template(
        "admin/imports_data.html",
        is_import_admin="Admin" in role_names,
        is_import_finance=bool(role_names & {"Finance Manager", "Finance Assistant"}),
    )


@admin_bp.route("/imports/grouped-franchises", methods=["GET", "POST"])
@login_required
def import_grouped_franchises():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the grouped franchises Excel file.", "danger")
            return redirect(url_for("admin.import_grouped_franchises"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_grouped_franchises"))

        processed = matched = created_users = 0
        grouped = []
        unmatched = []

        # Read the workbook in the exact format supplied by the user:
        #   Column A = main franchise user / main franchise
        #   Columns B:E = franchise users/franchises to link under column A
        #   Rows 2:14 contain the grouping data; row 1 is a heading row.
        #
        # The important rule is that Column A always remains the main franchise.
        # Columns B:E are linked to Column A, even if the same franchise used to
        # be linked somewhere else from a previous import.
        import_rows = []
        all_matched_franchise_ids = set()
        max_row = min(worksheet.max_row, 14)
        max_col = min(max(worksheet.max_column, 5), 5)
        for row_number in range(2, max_row + 1):
            main_name = clean_franchise_name(worksheet.cell(row_number, 1).value)
            if not main_name:
                continue

            linked_names = []
            seen_names = {normalize_franchise_key(main_name)}
            for col in range(2, max_col + 1):
                cell_value = worksheet.cell(row_number, col).value
                # A linked cell may contain one name or comma/semi-colon separated
                # names; split it safely while preserving the main in column A.
                for candidate in split_grouped_franchise_names(cell_value):
                    key = normalize_franchise_key(candidate)
                    if key and key not in seen_names:
                        linked_names.append(candidate)
                        seen_names.add(key)

            names = [main_name] + linked_names
            processed += 1
            resolved = []
            row_unmatched = []

            main_franchise = find_franchise_by_name(main_name)
            if not main_franchise:
                row_unmatched.append({
                    "row": row_number,
                    "name": main_name,
                    "reason": "Main franchise not found",
                    "linked": names,
                })
                unmatched.extend(row_unmatched)
                continue

            resolved.append(main_franchise)
            all_matched_franchise_ids.add(main_franchise.id)

            for name in linked_names:
                franchise = find_franchise_by_name(name)
                if franchise:
                    if franchise.id not in {item.id for item in resolved}:
                        resolved.append(franchise)
                        all_matched_franchise_ids.add(franchise.id)
                else:
                    row_unmatched.append({
                        "row": row_number,
                        "name": name,
                        "reason": "Linked franchise not found",
                        "linked": names,
                    })

            import_rows.append({
                "row": row_number,
                "names": names,
                "main": main_franchise,
                "franchises": resolved,
                "unmatched": row_unmatched,
            })
            unmatched.extend(row_unmatched)

        # Remove existing franchise-side links for every branch mentioned in the
        # sheet before adding the new links.  This is the important fix: old
        # groups such as Dobsonville -> Soweto cannot remain when the Excel row
        # says Soweto must be the main branch.  Admin/Finance users are not
        # touched.
        if all_matched_franchise_ids:
            franchise_side_roles = {"Franchise User", "Franchise Manager", "Read Only User"}
            franchise_side_user_ids = [
                row[0]
                for row in db.session.query(User.id)
                .join(User.roles)
                .filter(Role.name.in_(franchise_side_roles))
                .all()
            ]
            if franchise_side_user_ids:
                db.session.execute(
                    user_franchises.delete()
                    .where(user_franchises.c.franchise_id.in_(list(all_matched_franchise_ids)))
                    .where(user_franchises.c.user_id.in_(franchise_side_user_ids))
                )
                db.session.flush()

        for item in import_rows:
            main_franchise = item["main"]
            linked_franchises = item["franchises"]
            user, created = find_franchise_user_for_main_franchise(main_franchise)
            if created:
                created_users += 1
            set_primary_franchise_link(user, main_franchise, linked_franchises)
            matched += 1
            grouped.append({
                "row": item["row"],
                "main": main_franchise.business_name,
                "user": user.full_name,
                "email": user.email,
                "linked": [branch.business_name for branch in linked_franchises],
                "unmatched": [entry["name"] for entry in item["unmatched"]],
            })

        db.session.commit()
        log_action("Imports & Data", "Imported grouped franchises", f"Processed: {processed}, matched groups: {matched}, unmatched: {len(unmatched)}")
        flash(f"Grouped franchises imported. {matched} group(s) updated.", "success")
        return render_template(
            "admin/import_grouped_franchises.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            created_users=created_users,
            grouped=grouped,
            unmatched=unmatched,
        )

    return render_template(
        "admin/import_grouped_franchises.html",
        import_complete=False,
        processed=0,
        matched=0,
        created_users=0,
        grouped=[],
        unmatched=[],
    )


@admin_bp.route("/imports/contract-summary", methods=["GET", "POST"])
@login_required
def import_contract_summary():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the contract summary Excel file.", "danger")
            return redirect(url_for("admin.import_contract_summary"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_contract_summary"))

        groups = build_contract_summary_groups(worksheet)
        processed = matched = updated_scales = 0
        unmatched = []
        for group in groups:
            processed += 1
            franchise = find_franchise_by_name(group["name"])
            if not franchise:
                unmatched.append(group["name"])
                continue
            matched += 1
            rows = group["rows"]

            # Full refresh from the latest uploaded Excel file.
            # Dates are overwritten so changed agreement dates in the workbook update the Franchise Details page.
            start_date = newest_contract_date(worksheet.cell(r, 4).value for r in rows)
            end_date = newest_contract_date(worksheet.cell(r, 5).value for r in rows)
            franchise.agreement_start_date = start_date
            franchise.agreement_end_date = end_date
            set_auto_gross_method_from_agreement(franchise)

            ck_business_name = unique_join(worksheet.cell(r, 10).value for r in rows)
            ck_number = unique_join(worksheet.cell(r, 11).value for r in rows)
            pty_business_name = unique_join(worksheet.cell(r, 12).value for r in rows)
            pty_number = unique_join(worksheet.cell(r, 13).value for r in rows)
            franchisee = first_shareholder_name(worksheet.cell(r, 14).value for r in rows)

            # Full sync for contract master data: if the new file changed a value, replace it.
            franchise.ck_business_name = ck_business_name
            franchise.ck_number = ck_number
            franchise.pty_business_name = pty_business_name
            franchise.pty_number = pty_number
            if franchisee:
                first_name, surname = split_person_name(franchisee)
                franchise.franchisee_name = first_name
                franchise.franchisee_surname = surname
            else:
                franchise.franchisee_name = ""
                franchise.franchisee_surname = ""

            raw_scale_lines = []
            parsed_rows = []
            minimum = None
            for r in rows:
                parsed = parse_royalty_scale_line(worksheet.cell(r, 18).value)
                if not parsed:
                    continue
                raw = parsed.get("raw", "")
                if raw:
                    raw_scale_lines.append(raw)
                if parsed.get("minimum") is not None:
                    minimum = parsed["minimum"]
                    continue
                if "percentage" in parsed:
                    parsed_rows.append(parsed)

            # Full database sync from the latest uploaded Excel file.
            # This clears old scale rows first, then rebuilds them exactly from the new file.
            updated_scale_rows = sync_royalty_scales_from_contract_file(franchise, parsed_rows, raw_scale_lines, minimum)
            if updated_scale_rows:
                updated_scales += 1

        # Recalculate existing monthly figures because the agreement date controls
        # whether the franchise uses Gross = New Gross Method or Gross = Old.
        from app.monthly.routes import recalculate_monthly_figure
        for franchise in Franchise.query.all():
            for figure in MonthlyFigure.query.filter_by(franchise_id=franchise.id).all():
                recalculate_monthly_figure(figure)

        db.session.commit()
        log_action("Imports & Data", "Imported contract summary", f"Processed: {processed}, matched: {matched}, unmatched: {len(unmatched)}")
        flash(f"Contract summary import complete. {processed} franchises processed, {matched} matched, {len(unmatched)} unmatched, {updated_scales} royalty scales updated.", "success")
        return render_template(
            "admin/import_contract_summary.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            unmatched=unmatched,
            updated_scales=updated_scales,
        )

    return render_template("admin/import_contract_summary.html", import_complete=False, processed=0, matched=0, unmatched=[], updated_scales=0)





def apply_contact_data_to_franchise(franchise, contact_name="", office_number="", email="", address=""):
    """Apply imported/manual contact-list values to one Franchise record."""
    contact_name = normalize_contact_value(contact_name)
    office_number = normalize_contact_value(office_number)
    email = normalize_contact_value(email)
    address = normalize_contact_value(address)

    # Full sync: every new upload overwrites the old database values.
    franchise.office_number = office_number
    franchise.after_hours_number = office_number
    franchise.franchisee_cell = office_number
    franchise.franchisee_email = email
    franchise.public_email = email
    franchise.office_address = address
    first_name, surname = split_person_name(contact_name)
    franchise.franchisee_name = first_name or ""
    franchise.franchisee_surname = surname or ""

@admin_bp.route("/imports/contact-list", methods=["GET", "POST"])
@login_required
def import_contact_list():
    if not is_current_user_admin():
        abort(403)
    if request.method == "POST":
        uploaded_file = request.files.get("excel_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("Please upload the Martins Funerals contact list Excel file.", "danger")
            return redirect(url_for("admin.import_contact_list"))
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            flash(f"Could not read Excel file: {exc}", "danger")
            return redirect(url_for("admin.import_contact_list"))

        processed = matched = 0
        unmatched = []
        updated = []
        # The uploaded contact list currently has headers in row 1 and data from row 2.
        # Process all populated rows, not only the first 76, because some files include
        # more than 76 outlets/branches.
        last_row = worksheet.max_row
        for row_number in range(2, last_row + 1):
            branch = normalize_contact_value(worksheet.cell(row_number, 2).value)
            outlet = normalize_contact_value(worksheet.cell(row_number, 3).value)
            if not branch and not outlet:
                continue
            processed += 1
            candidates = contact_candidate_names(branch, outlet)
            franchise, matched_name = find_franchise_by_candidates(candidates)
            display_name = " / ".join([item for item in [branch, outlet] if item])
            if not franchise:
                unmatched.append({
                    "row": row_number,
                    "name": display_name or f"Row {row_number}",
                    "branch": branch,
                    "outlet": outlet,
                    "contact_name": normalize_contact_value(worksheet.cell(row_number, 4).value),
                    "office_number": normalize_contact_value(worksheet.cell(row_number, 6).value),
                    "email": normalize_contact_value(worksheet.cell(row_number, 7).value),
                    "address": normalize_contact_value(worksheet.cell(row_number, 8).value),
                    "tried": ", ".join(candidates),
                })
                continue

            matched += 1
            contact_name = normalize_contact_value(worksheet.cell(row_number, 4).value)
            # User-approved mapping for this file:
            # B/C = franchise name/outlet, F = office number, 24-hour number and cell number,
            # G = franchisee/public email, H = office address.
            office_number = normalize_contact_value(worksheet.cell(row_number, 6).value)
            email = normalize_contact_value(worksheet.cell(row_number, 7).value)
            address = normalize_contact_value(worksheet.cell(row_number, 8).value)

            # Store the uploaded contact information on the Franchise record so every
            # linked franchise user sees it on the Franchise Details page.
            apply_contact_data_to_franchise(franchise, contact_name, office_number, email, address)

            updated.append({
                "row": row_number,
                "spreadsheet_name": display_name,
                "matched_franchise": franchise.business_name,
                "matched_by": matched_name,
                "email": email,
                "office_number": office_number,
            })

        db.session.commit()
        log_action("Imports & Data", "Imported contact list", f"Processed: {processed}, matched: {matched}, unmatched: {len(unmatched)}")
        flash(f"Contact list import complete. {processed} rows processed, {matched} matched, {len(unmatched)} unmatched.", "success")
        return render_template(
            "admin/import_contact_list.html",
            import_complete=True,
            processed=processed,
            matched=matched,
            unmatched=unmatched,
            updated=updated,
        )

    return render_template("admin/import_contact_list.html", import_complete=False, processed=0, matched=0, unmatched=[], updated=[])



@admin_bp.route("/imports/contact-list/manual", methods=["GET", "POST"])
@login_required
def manual_allocate_contact_list_row():
    if not is_current_user_admin():
        abort(403)
    franchises = Franchise.query.order_by(Franchise.business_name).all()
    if request.method == "POST":
        franchise_id = request.form.get("franchise_id", type=int)
        franchise = Franchise.query.get(franchise_id) if franchise_id else None
        if not franchise:
            flash("Please select the correct franchise before saving.", "danger")
            return redirect(url_for("admin.import_contact_list"))

        contact_name = request.form.get("contact_name", "")
        office_number = request.form.get("office_number", "")
        email = request.form.get("email", "")
        address = request.form.get("address", "")
        spreadsheet_name = request.form.get("spreadsheet_name", "")
        row_number = request.form.get("row_number", "")

        apply_contact_data_to_franchise(franchise, contact_name, office_number, email, address)
        db.session.commit()
        log_action(
            "Imports & Data",
            "Manually allocated contact-list row",
            f"Row: {row_number}; Spreadsheet: {spreadsheet_name}; Franchise: {franchise.business_name}",
        )
        flash(f"Contact-list row allocated to {franchise.business_name} and saved to Franchise Details.", "success")
        return redirect(url_for("admin.import_contact_list"))

    row_data = {
        "row_number": request.args.get("row", ""),
        "spreadsheet_name": request.args.get("name", ""),
        "branch": request.args.get("branch", ""),
        "outlet": request.args.get("outlet", ""),
        "contact_name": request.args.get("contact_name", ""),
        "office_number": request.args.get("office_number", ""),
        "email": request.args.get("email", ""),
        "address": request.args.get("address", ""),
        "tried": request.args.get("tried", ""),
    }
    return render_template("admin/manual_allocate_contact.html", franchises=franchises, row_data=row_data)

@admin_bp.route("/roles")
@login_required
@permission_required("user_roles:view")
def roles():
    roles = Role.query.order_by(Role.name).all()
    return render_template("admin/roles.html", roles=roles)


@admin_bp.route("/roles/<int:role_id>", methods=["GET", "POST"])
@login_required
@permission_required("user_roles:edit")
def edit_role(role_id):
    role = Role.query.get_or_404(role_id)
    permissions = Permission.query.order_by(Permission.sort_order).all()
    grouped = defaultdict(list)
    for permission in permissions:
        grouped[permission.module].append(permission)

    if request.method == "POST":
        selected_ids = [int(item) for item in request.form.getlist("permission_ids")]
        role.permissions = Permission.query.filter(Permission.id.in_(selected_ids)).all() if selected_ids else []
        log_action("User Roles", "Updated role permissions", f"Role: {role.name}")
        db.session.commit()
        flash(f"Permissions updated for {role.name}.", "success")
        return redirect(url_for("admin.roles"))

    selected = {permission.id for permission in role.permissions}
    return render_template("admin/edit_role.html", role=role, grouped_permissions=grouped, actions=ACTIONS, selected=selected)


@admin_bp.route("/roles/new", methods=["GET", "POST"])
@login_required
@permission_required("user_roles:add")
def new_role():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Role name is required.", "danger")
            return render_template("admin/new_role.html")
        if Role.query.filter_by(name=name).first():
            flash("A role with this name already exists.", "warning")
            return render_template("admin/new_role.html")
        role = Role(name=name, description=description)
        db.session.add(role)
        log_action("User Roles", "Created role", f"Role: {role.name}")
        db.session.commit()
        flash("Role created. You can now tick its permissions.", "success")
        return redirect(url_for("admin.edit_role", role_id=role.id))
    return render_template("admin/new_role.html")


@admin_bp.route("/audit-logs")
@login_required
@permission_required("audit_logs:view")
def audit_logs():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(300).all()
    return render_template("admin/audit_logs.html", logs=logs)

@admin_bp.route("/users/<int:user_id>/activate", methods=["POST"])
@login_required
@permission_required("users:edit")
def activate_user(user_id):
    user = User.query.get_or_404(user_id)

    user.is_active = True

    db.session.commit()

    flash(f"{user.name} activated.", "success")

    return redirect(url_for("admin.users"))

@admin_bp.route("/users/<int:user_id>/franchises", methods=["POST"])
@login_required
@permission_required("users:edit")
def assign_user_franchises(user_id):
    if not can_assign_franchise_links():
        flash("Your role does not have permission to link franchise users to franchises.", "danger")
        return redirect(url_for("admin.users"))

    user = User.query.get_or_404(user_id)
    if is_admin_side_user(user):
        user.assigned_franchises = []
        db.session.commit()
        flash("Martins Funerals South Africa/admin-side users are not linked to franchises here.", "warning")
        return redirect(url_for("admin.users"))

    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]
    user.assigned_franchises = Franchise.query.filter(Franchise.id.in_(franchise_ids)).all() if franchise_ids else []
    log_action("Users", "Updated user franchise access", f"User: {user.full_name}")
    db.session.commit()
    flash(f"Franchise access updated for {user.full_name}.", "success")
    return redirect(url_for("admin.users"))


@admin_bp.route("/users/clean-finance-admin-users", methods=["POST"])
@login_required
@permission_required("users:edit")
def clean_finance_admin_users():
    changed = tidy_finance_admin_users()
    log_action("Users", "Cleaned finance/admin-side users", f"Changes: {changed}")
    db.session.commit()
    flash("Renette, Lowhaan and Deon were cleaned as Martins Funerals South Africa/admin-side users with no franchise links.", "success")
    return redirect(url_for("admin.users"))

@admin_bp.route("/users/clear-franchise-user-links", methods=["POST"])
@login_required
@permission_required("users:edit")
def clear_franchise_user_links():
    if not can_assign_franchise_links():
        flash("Your role does not have permission to clear franchise user links.", "danger")
        return redirect(url_for("admin.users"))

    target_role_names = {"Franchise User", "Franchise Manager", "Regional Manager"}
    protected_role_names = {"Admin", "Finance Manager", "Finance Assistant"}
    cleared = 0

    users = User.query.all()
    for user in users:
        role_names = {role.name for role in user.roles}
        if role_names & target_role_names and not (role_names & protected_role_names):
            if user.assigned_franchises:
                user.assigned_franchises = []
                cleared += 1

    log_action("Users", "Cleared franchise user links", f"Users cleared: {cleared}")
    db.session.commit()
    flash(f"Cleared linked franchises from {cleared} franchise user(s). Finance and admin users were not changed.", "success")
    return redirect(url_for("admin.users"))

# ---------------------------------------------------------------------------
# Admin oversight for employee users created by franchise users
# ---------------------------------------------------------------------------

def is_franchise_employee_user(user):
    """Return True for any employee account that belongs under a franchise user.

    Older records were not always saved with parent_franchise_user_id, so Admin
    must also treat the franchise-side employee roles as employee accounts. This
    keeps Manager, Employee and Agent accounts visible/editable in Admin >
    Employees.
    """
    return (
        user.has_role("Franchise Manager")
        or user.has_role("Franchise Employee")
        or user.has_role("Franchise Agent")
        or bool(getattr(user, "parent_franchise_user_id", None))
        or bool(getattr(user, "created_by_user_id", None))
    )


@admin_bp.route("/franchise-employees")
@login_required
@permission_required("users:view")
def franchise_employees():
    employees = [user for user in User.query.order_by(User.name, User.surname).all() if is_franchise_employee_user(user)]
    owner_ids = [employee.parent_franchise_user_id for employee in employees if employee.parent_franchise_user_id]
    owners = {user.id: user for user in User.query.filter(User.id.in_(owner_ids)).all()} if owner_ids else {}
    now = datetime.utcnow()
    franchise_owners = active_recent_franchise_owner_users(now.month, now.year)
    franchises = Franchise.query.filter(Franchise.is_performance_active == True).order_by(Franchise.business_name).all()
    employee_roles = Role.query.filter(Role.name.in_(["Franchise Manager", "Franchise Employee", "Franchise Agent"])).order_by(Role.name).all()
    return render_template(
        "admin/franchise_employees.html",
        employees=employees,
        owners=owners,
        franchise_owners=franchise_owners,
        franchises=franchises,
        employee_roles=employee_roles,
    )


@admin_bp.route("/franchise-employees/create", methods=["POST"])
@login_required
@permission_required("users:add")
def create_franchise_employee_admin():
    name = request.form.get("name", "").strip()
    surname = request.form.get("surname", "").strip()
    email = request.form.get("email", "").strip().lower()
    password = request.form.get("password", "").strip()
    role_id = request.form.get("role_id", type=int)
    franchise_id = request.form.get("franchise_id", type=int)
    owner_id = request.form.get("parent_franchise_user_id", type=int)

    if not name or not surname or not email or not password or not role_id or not franchise_id:
        flash("Name, surname, email, password, role and franchise are required.", "danger")
        return redirect(url_for("admin.franchise_employees"))
    if User.query.filter(db.func.lower(User.email) == email).first():
        flash("A user with that email address already exists.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    selected_role = Role.query.get(role_id)
    if not selected_role or selected_role.name not in {"Franchise Manager", "Franchise Employee", "Franchise Agent"}:
        flash("Please select Manager, Employee or Agent.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    franchise = Franchise.query.get_or_404(franchise_id)
    owner = User.query.get(owner_id) if owner_id else None
    if owner and (not owner.has_role("Franchise User") or not franchise_user_has_recent_kpi_data(owner)):
        owner = None
    if not owner:
        owner = next((candidate for candidate in active_recent_franchise_owner_users() if franchise in candidate.assigned_franchises), None)

    user = User(
        name=name,
        surname=surname,
        email=email,
        is_active=True,
        is_active_account=True,
        parent_franchise_user_id=owner.id if owner else None,
        created_by_user_id=current_user.id,
    )
    user.set_password(password)
    user.roles.append(selected_role)
    user.assigned_franchises.append(franchise)
    db.session.add(user)
    log_action("Franchise Employees", "Admin created franchise employee user", f"Employee: {email}; Franchise: {franchise.business_name}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was created.", "success")
    return redirect(url_for("admin.franchise_employees"))


@admin_bp.route("/franchise-employees/<int:user_id>/update", methods=["POST"])
@login_required
@permission_required("users:edit")
def update_franchise_employee(user_id):
    user = User.query.get_or_404(user_id)
    if not is_franchise_employee_user(user):
        flash("This user is not a franchise employee user.", "danger")
        return redirect(url_for("admin.franchise_employees"))

    user.name = request.form.get("name", user.name).strip() or user.name
    user.surname = request.form.get("surname", user.surname).strip() or user.surname

    role_id = request.form.get("role_id", type=int)
    if role_id:
        selected_role = Role.query.get(role_id)
        if not selected_role or selected_role.name not in {"Franchise Manager", "Franchise Employee", "Franchise Agent"}:
            flash("Please select Manager, Employee or Agent for franchise employees.", "danger")
            return redirect(url_for("admin.franchise_employees"))
        user.roles = [selected_role]

    password = request.form.get("password", "").strip()
    if password:
        user.set_password(password)
    user.is_active = request.form.get("is_active") == "1"

    franchise_ids = [int(item) for item in request.form.getlist("franchise_ids")]
    user.assigned_franchises = Franchise.query.filter(Franchise.id.in_(franchise_ids)).all() if franchise_ids else []

    owner_id = request.form.get("parent_franchise_user_id", type=int)
    if owner_id:
        owner = User.query.get(owner_id)
        if owner and owner != user:
            user.parent_franchise_user_id = owner.id

    log_action("Franchise Employees", "Admin updated franchise employee user", f"Employee: {user.email}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was updated.", "success")
    return redirect(url_for("admin.franchise_employees"))


@admin_bp.route("/franchise-employees/<int:user_id>/delete", methods=["POST"])
@login_required
@permission_required("users:delete")
def delete_franchise_employee(user_id):
    user = User.query.get_or_404(user_id)
    if not is_franchise_employee_user(user):
        flash("This user is not a franchise employee user.", "danger")
        return redirect(url_for("admin.franchise_employees"))
    user.is_active = False
    user.is_active_account = False
    user.deactivated_at = datetime.utcnow()
    user.deactivation_reason = "Deactivated by Admin"
    log_action("Franchise Employees", "Admin deactivated franchise employee user", f"Employee: {user.email}")
    db.session.commit()
    flash(f"Employee user {user.full_name} was deactivated.", "success")
    return redirect(url_for("admin.franchise_employees"))
