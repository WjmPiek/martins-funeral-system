from collections import Counter
from datetime import datetime
from functools import wraps

from flask import Blueprint, abort, current_app, jsonify, render_template, request, redirect, url_for, flash
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app.audit import log_action
from app.extensions import db
from app.franchise_context import get_selected_franchise, is_franchise_view_mode
from app.models import Franchise, HeatmapRecord

heatmap_bp = Blueprint("heatmap", __name__, url_prefix="/heat-map")

PROVINCES = [
    "Eastern Cape", "Free State", "Gauteng", "KwaZulu-Natal", "Limpopo",
    "Mpumalanga", "North West", "Northern Cape", "Western Cape",
]


def permission_required(code):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.has_permission(code):
                abort(403)
            return func(*args, **kwargs)
        return wrapper
    return decorator


def can_modify_heatmap():
    return any(current_user.has_permission(code) for code in [
        "heat_map:add", "heat_map:edit", "heat_map:import", "heat_map:manage"
    ])


def accessible_franchises():
    return current_user.accessible_franchises()


def accessible_franchise_ids():
    return [franchise.id for franchise in accessible_franchises()]


def selected_or_requested_franchise_id():
    requested = request.values.get("franchise_id", type=int)
    selected = get_selected_franchise() if is_franchise_view_mode() else None
    franchise_id = requested or (selected.id if selected else None)
    allowed = set(accessible_franchise_ids())
    if franchise_id and franchise_id not in allowed:
        abort(403)
    return franchise_id


def scoped_query():
    query = HeatmapRecord.query
    allowed = accessible_franchise_ids()
    if not (current_user.has_permission("franchise_management:view") or current_user.has_permission("franchise_management:manage")):
        if not allowed:
            return query.filter(False)
        query = query.filter(HeatmapRecord.franchise_id.in_(allowed))
    franchise_id = request.args.get("franchise_id", type=int)
    if franchise_id:
        if franchise_id not in allowed and not current_user.has_permission("franchise_management:view"):
            abort(403)
        query = query.filter_by(franchise_id=franchise_id)
    return query


def clean(value):
    return " ".join(str(value or "").strip().split())


def number(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def header_map(row):
    result = {}
    for index, value in enumerate(row, start=1):
        key = clean(value).lower().replace("_", " ").replace("-", " ")
        if key:
            result[key] = index
    return result


def cell(ws, row, headers, *names):
    for name in names:
        idx = headers.get(name)
        if idx:
            return ws.cell(row, idx).value
    return ""


def build_full_address(address, city, province, country):
    return ", ".join(part for part in [clean(address), clean(city), clean(province), clean(country) or "South Africa"] if part)


def parse_heatmap_excel(file_storage, source_filename, franchise_id):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = header_map([cell.value for cell in ws[1]])
    if not headers:
        raise ValueError("The uploaded Excel file does not contain a header row.")

    has_relation_column = "relation" in headers
    records = []
    skipped_non_mem = 0
    skipped_blank = 0

    for row in range(2, ws.max_row + 1):
        mf_file = clean(cell(ws, row, headers, "mf file", "mf_file", "file", "policy number"))
        deceased_name = clean(cell(ws, row, headers, "deceased name", "name"))
        deceased_surname = clean(cell(ws, row, headers, "deceased surname", "surname"))
        address = clean(cell(ws, row, headers, "address", "street address", "residential address"))
        city = clean(cell(ws, row, headers, "city", "town", "town city"))
        province = clean(cell(ws, row, headers, "province"))
        country = clean(cell(ws, row, headers, "country")) or "South Africa"
        full_address = clean(cell(ws, row, headers, "full address", "fulladdress")) or build_full_address(address, city, province, country)
        relation = clean(cell(ws, row, headers, "relation"))

        if has_relation_column and relation.upper() != "MEM":
            skipped_non_mem += 1
            continue
        if not any([mf_file, deceased_name, deceased_surname, address, city, province, full_address]):
            skipped_blank += 1
            continue

        records.append(HeatmapRecord(
            franchise_id=franchise_id,
            mf_file=mf_file,
            deceased_name=deceased_name,
            deceased_surname=deceased_surname,
            dod=clean(cell(ws, row, headers, "dod", "date of death")),
            address=address,
            city=city,
            province=province,
            country=country,
            full_address=full_address,
            latitude=number(cell(ws, row, headers, "latitude", "lat")),
            longitude=number(cell(ws, row, headers, "longitude", "lng", "lon")),
            weight=number(cell(ws, row, headers, "weight")) or 1,
            next_of_kin_name=clean(cell(ws, row, headers, "next of kin name", "nok name")),
            next_of_kin_surname=clean(cell(ws, row, headers, "next of kin surname", "nok surname")),
            relationship=clean(cell(ws, row, headers, "relationship")),
            relation=relation,
            contact_number=clean(cell(ws, row, headers, "contact number", "cell number", "phone")),
            source_filename=source_filename,
            created_by_id=current_user.id,
        ))
    return records, skipped_non_mem, skipped_blank


@heatmap_bp.route("/")
@login_required
@permission_required("heat_map:view")
def index():
    franchises = accessible_franchises()
    selected = get_selected_franchise() if is_franchise_view_mode() else None
    return render_template(
        "heatmap/index.html",
        franchises=franchises,
        selected_franchise=selected,
        provinces=PROVINCES,
        can_modify_heatmap=can_modify_heatmap(),
        google_maps_api_key=current_app.config.get("GOOGLE_MAPS_API_KEY", ""),
    )


@heatmap_bp.route("/data")
@login_required
@permission_required("heat_map:view")
def data():
    records = scoped_query().order_by(HeatmapRecord.city.asc(), HeatmapRecord.mf_file.asc()).all()
    province_counts = Counter(record.province for record in records if record.province)
    city_counts = Counter(record.city for record in records if record.city)
    mapped = sum(1 for record in records if record.latitude is not None and record.longitude is not None)
    return jsonify({
        "records": [record.to_dict() for record in records],
        "summary": {
            "total": len(records),
            "mapped": mapped,
            "unmapped": len(records) - mapped,
            "province": dict(province_counts),
            "cities": dict(city_counts.most_common(10)),
        }
    })


@heatmap_bp.route("/import", methods=["POST"])
@login_required
@permission_required("heat_map:import")
def import_excel():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose an Excel file to import.", "danger")
        return redirect(url_for("heatmap.index"))
    filename = file.filename
    franchise_id = selected_or_requested_franchise_id()
    if not franchise_id and accessible_franchise_ids():
        franchise_id = request.form.get("franchise_id", type=int) or None
    if franchise_id and franchise_id not in accessible_franchise_ids() and not current_user.has_permission("franchise_management:view"):
        abort(403)
    try:
        records, skipped_non_mem, skipped_blank = parse_heatmap_excel(file, filename, franchise_id)
        if not records:
            flash("No valid heat map rows were found in the uploaded file.", "warning")
            return redirect(url_for("heatmap.index"))
        db.session.add_all(records)
        db.session.commit()
        detail = f"Imported {len(records)} heat map records from {filename}. Skipped non-MEM rows: {skipped_non_mem}; blank rows: {skipped_blank}."
        log_action("Heat Map", "Imported heat map records", detail)
        flash(detail, "success")
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Heat map import failed: %s", exc)
        flash(str(exc) or "Heat map import failed.", "danger")
    return redirect(url_for("heatmap.index"))


@heatmap_bp.route("/record", methods=["POST"])
@login_required
def save_record():
    if not can_modify_heatmap():
        abort(403)
    payload = request.get_json(force=True)
    record_id = payload.get("id")
    record = HeatmapRecord.query.get(record_id) if record_id else HeatmapRecord(created_by_id=current_user.id)
    if not record:
        abort(404)
    if record.franchise_id and record.franchise_id not in accessible_franchise_ids() and not current_user.has_permission("franchise_management:view"):
        abort(403)
    franchise_id = payload.get("franchiseId") or selected_or_requested_franchise_id()
    if franchise_id:
        franchise_id = int(franchise_id)
        if franchise_id not in accessible_franchise_ids() and not current_user.has_permission("franchise_management:view"):
            abort(403)
    record.franchise_id = franchise_id
    mapping = {
        "mf_file": "mfFile", "deceased_name": "deceasedName", "deceased_surname": "deceasedSurname",
        "dod": "dod", "address": "address", "city": "city", "province": "province", "country": "country",
        "full_address": "fullAddress", "next_of_kin_name": "nextOfKinName", "next_of_kin_surname": "nextOfKinSurname",
        "relationship": "relationship", "relation": "relation", "contact_number": "contactNumber",
    }
    for attr, key in mapping.items():
        setattr(record, attr, clean(payload.get(key)))
    record.latitude = number(payload.get("latitude"))
    record.longitude = number(payload.get("longitude"))
    record.weight = number(payload.get("weight")) or 1
    if not record.full_address:
        record.full_address = build_full_address(record.address, record.city, record.province, record.country)
    db.session.add(record)
    db.session.commit()
    log_action("Heat Map", "Saved heat map record", record.mf_file or record.full_address or str(record.id))
    return jsonify({"record": record.to_dict()})


@heatmap_bp.route("/record/<int:record_id>/delete", methods=["POST"])
@login_required
def delete_record(record_id):
    if not can_modify_heatmap():
        abort(403)
    record = HeatmapRecord.query.get_or_404(record_id)
    if record.franchise_id and record.franchise_id not in accessible_franchise_ids() and not current_user.has_permission("franchise_management:view"):
        abort(403)
    db.session.delete(record)
    db.session.commit()
    log_action("Heat Map", "Deleted heat map record", record.mf_file or record.full_address or str(record_id))
    return jsonify({"message": "Deleted"})
