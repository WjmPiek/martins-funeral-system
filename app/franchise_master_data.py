"""Franchise Master data integrity and workbook import/export helpers.

This module keeps franchise master data in one place.  It is deliberately
conservative: it matches existing franchise records by ID first, then code,
then normalized business name.  It never creates duplicate franchises unless
explicitly extended later.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import io
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from flask import send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from app import db
from app.models import Franchise, MonthlyFigure, RoyaltyCalculationSnapshot, RoyaltyScale

PROVINCES = [
    "Eastern Cape", "Free State", "Gauteng", "KwaZulu-Natal", "Limpopo",
    "Mpumalanga", "North West", "Northern Cape", "Western Cape", "Unassigned",
]

ROYALTY_SCALE_COUNT = 7

MASTER_HEADERS = [
    "Franchise ID", "Business Name", "Franchise Code", "Province", "Region", "District", "Municipality",
    "Office Address", "Office Number", "After Hours Number", "Franchisee Name", "Franchisee Surname",
    "Franchisee Cell", "Franchisee Email", "Public Email", "Agreement Start Date", "Agreement End Date",
    "Royalty Method", "Minimum Royalty Amount",
]
for idx in range(1, ROYALTY_SCALE_COUNT + 1):
    MASTER_HEADERS += [f"Scale {idx} From", f"Scale {idx} To", f"Scale {idx} %"]

PROVINCE_TERMS = {
    "Gauteng": ["alberton", "benoni", "boksburg", "brakpan", "springs", "edenvale", "germiston", "katlehong", "vosloorus", "tsakane", "thokoza", "tokoza", "tembisa", "midrand", "pretoria", "soshanguve", "sochanguve", "mamelodi", "atteridgeville", "centurion", "hammanskraal", "vereeniging", "vanderbijlpark", "meyerton", "sebokeng", "orange farm", "lenasia", "three rivers", "florida", "fountainbleau", "carletonville", "randfontein", "krugersdorp", "roodepoort", "soweto", "sandton", "randburg", "johannesburg"],
    "Western Cape": ["cape town", "brackenfell", "bellville", "paarl", "parow", "kraaifontein", "kuils river", "durbanville", "table view", "stellenbosch", "strand", "somerset west", "worcester", "george", "mossel bay", "mosselbaai", "oudtshoorn", "knysna", "plettenberg", "beaufort west", "malmesbury", "vredenburg", "saldanha", "hermanus", "caledon", "robertson", "wellington"],
    "KwaZulu-Natal": ["durban", "pinetown", "phoenix", "umlazi", "umhlanga", "chatsworth", "isipingo", "kwamashu", "verulam", "tongaat", "pietermaritzburg", "empangeni", "richards bay", "ladysmith", "newcastle", "estcourt", "kokstad", "port shepstone", "margate", "vryheid", "eshowe", "ulundi", "stanger", "kwadukuza", "ballito"],
    "Eastern Cape": ["gqeberha", "port elizabeth", "east london", "mthatha", "umtata", "queenstown", "komani", "jeffreys bay", "jeffreys baai", "jeffreysbaai", "humansdorp", "uitenhage", "kariega", "grahamstown", "makhanda", "cradock", "graaff", "butterworth"],
    "Limpopo": ["polokwane", "pietersburg", "tzaneen", "mokopane", "potgietersrus", "mookgophong", "mookgopong", "modimolle", "nylstroom", "bela-bela", "belabela", "thohoyandou", "louis trichardt", "makhado", "giyani", "phalaborwa", "lephalale", "ellisras", "musina", "seshego"],
    "Mpumalanga": ["mbombela", "nelspruit", "witbank", "emalahleni", "middelburg", "secunda", "evander", "bethal", "ermelo", "piet retief", "barberton", "lydenburg", "white river", "hazyview", "komatipoort", "standerton", "volksrust", "delmas", "kriel"],
    "North West": ["rustenburg", "klerksdorp", "potchefstroom", "mahikeng", "mafikeng", "brits", "lichtenburg", "vryburg", "orkney", "stilfontein", "hartbeespoort", "zeerust", "taung", "wolmaransstad", "christiana"],
    "Free State": ["bloemfontein", "welkom", "bethlehem", "kroonstad", "sasolburg", "sasolsburg", "virginia", "harrismith", "parys", "ficksburg", "phuthaditjhaba", "botshabelo", "ladybrand", "senekal", "heilbron"],
    "Northern Cape": ["kimberley", "upington", "kuruman", "springbok", "de aar", "postmasburg", "kathu", "hartswater", "colesberg", "calvinia", "prieska", "douglas", "jan kempdorp", "barkly west", "warrenton", "hopetown"],
}


def normalize_key(value: Any) -> str:
    value = str(value or "").lower().strip()
    value = re.sub(r"\(f\)", "", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def infer_province(*values: Any) -> str:
    haystack = normalize_key(" ".join(str(v or "") for v in values))
    if not haystack:
        return ""
    for province, terms in PROVINCE_TERMS.items():
        for term in terms:
            if normalize_key(term) in haystack:
                return province
    return ""


def parse_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def decimal_or_zero(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, str):
        value = value.replace("R", "").replace("%", "").replace(",", "").strip()
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def get_latest_period() -> Tuple[int, int]:
    row = db.session.query(MonthlyFigure.year, MonthlyFigure.month).order_by(MonthlyFigure.year.desc(), MonthlyFigure.month.desc()).first()
    if row:
        return int(row.year), int(row.month)
    today = date.today()
    return today.year, today.month


def franchise_lookup() -> Tuple[Dict[int, Franchise], Dict[str, Franchise], Dict[str, Franchise]]:
    by_id, by_code, by_name = {}, {}, {}
    for franchise in Franchise.query.all():
        by_id[franchise.id] = franchise
        if franchise.franchise_code:
            by_code[normalize_key(franchise.franchise_code)] = franchise
        by_name[normalize_key(franchise.business_name)] = franchise
    return by_id, by_code, by_name


def find_franchise(row: Dict[str, Any], lookups=None) -> Optional[Franchise]:
    by_id, by_code, by_name = lookups or franchise_lookup()
    raw_id = row.get("Franchise ID")
    try:
        fid = int(raw_id) if raw_id not in (None, "") else None
    except Exception:
        fid = None
    if fid and fid in by_id:
        return by_id[fid]
    code = normalize_key(row.get("Franchise Code"))
    if code and code in by_code:
        return by_code[code]
    name = normalize_key(row.get("Business Name"))
    return by_name.get(name)


def data_integrity_rows() -> List[Dict[str, Any]]:
    rows = []
    latest_year, latest_month = get_latest_period()
    latest_review = {
        r.franchise_id: r for r in RoyaltyCalculationSnapshot.query.filter_by(year=latest_year, month=latest_month).all()
    }
    for f in Franchise.query.order_by(Franchise.business_name).all():
        issues = []
        if not f.business_name:
            issues.append("Business name missing")
        if not (getattr(f, "province", "") or "").strip() or getattr(f, "province", "") == "Unassigned":
            issues.append("Province not assigned")
        if not (getattr(f, "region", "") or "").strip():
            issues.append("Region not assigned")
        if not f.office_address:
            issues.append("Office address missing")
        if not (f.office_number or f.after_hours_number):
            issues.append("Contact number missing")
        if not f.agreement_start_date:
            issues.append("Agreement start date missing")
        scale_count = RoyaltyScale.query.filter_by(franchise_id=f.id).count()
        if scale_count == 0:
            issues.append("Royalty scale missing")
        review = latest_review.get(f.id)
        if review and review.status == "needs_review":
            issues.append(review.reason or "Royalty needs review")
        rows.append({
            "id": f.id,
            "business_name": f.business_name,
            "province": getattr(f, "province", "") or "Unassigned",
            "region": getattr(f, "region", "") or "",
            "agreement_start_date": f.agreement_start_date,
            "scale_count": scale_count,
            "issue_count": len(issues),
            "issues": issues,
            "status": "Ready" if not issues else "Needs Review",
        })
    return rows


def assign_regions_from_existing_data(commit: bool = True) -> Dict[str, int]:
    updated = 0
    unassigned = 0
    for f in Franchise.query.order_by(Franchise.business_name).all():
        detected = infer_province(f.business_name, f.office_address, f.franchise_code)
        if detected:
            if (getattr(f, "province", "") or "") != detected:
                f.province = detected
                updated += 1
            if not getattr(f, "region", ""):
                f.region = detected
        else:
            if not getattr(f, "province", ""):
                f.province = "Unassigned"
            unassigned += 1
    if commit:
        db.session.commit()
    return {"updated": updated, "unassigned": unassigned}


def build_franchise_master_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Franchise Master"
    ws.append(MASTER_HEADERS)
    franchises = Franchise.query.order_by(Franchise.business_name).all()
    for f in franchises:
        scales = RoyaltyScale.query.filter_by(franchise_id=f.id).order_by(RoyaltyScale.row_number.asc()).all()
        row = [
            f.id, f.business_name, f.franchise_code, getattr(f, "province", "") or "", getattr(f, "region", "") or "",
            getattr(f, "district", "") or "", getattr(f, "municipality", "") or "", f.office_address, f.office_number,
            f.after_hours_number, f.franchisee_name, f.franchisee_surname, f.franchisee_cell, f.franchisee_email,
            f.public_email, f.agreement_start_date, f.agreement_end_date, f.royalty_gross_method, float(f.minimum_royalty_amount or 0),
        ]
        for i in range(ROYALTY_SCALE_COUNT):
            s = scales[i] if i < len(scales) else None
            row.extend([float(s.amount_from or 0) if s else None, float(s.amount_to or 0) if s else None, float(s.percentage or 0) if s else None])
        ws.append(row)

    _style_master_sheet(ws)
    _add_validation_sheet(wb, len(franchises) + 100)
    _add_instructions_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_master_sheet(ws):
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="153D2A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2D5")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    widths = {
        1: 12, 2: 28, 3: 16, 4: 18, 5: 18, 6: 18, 7: 20, 8: 36, 9: 18, 10: 18,
        11: 18, 12: 18, 13: 18, 14: 28, 15: 28, 16: 16, 17: 16, 18: 14, 19: 16,
    }
    for col_idx in range(1, len(MASTER_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in [16, 17]:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "yyyy-mm-dd"
    for col in range(19, len(MASTER_HEADERS) + 1):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "#,##0.00"


def _add_validation_sheet(wb, max_rows: int):
    ws = wb.create_sheet("Lists")
    ws.sheet_state = "hidden"
    ws.append(["Province", "Royalty Method"])
    for idx, province in enumerate(PROVINCES, start=2):
        ws.cell(idx, 1).value = province
    for idx, method in enumerate(["old", "new"], start=2):
        ws.cell(idx, 2).value = method
    master = wb["Franchise Master"]
    province_val = DataValidation(type="list", formula1="=Lists!$A$2:$A$10", allow_blank=True)
    method_val = DataValidation(type="list", formula1="=Lists!$B$2:$B$3", allow_blank=True)
    master.add_data_validation(province_val)
    master.add_data_validation(method_val)
    province_val.add(f"D2:D{max_rows}")
    method_val.add(f"R2:R{max_rows}")


def _add_instructions_sheet(wb):
    ws = wb.create_sheet("Instructions", 0)
    lines = [
        ["Franchise Master Update Template"],
        ["1. Do not change Franchise ID unless you are intentionally matching a different existing franchise."],
        ["2. Update business details, contact numbers, province/region, agreement dates and royalty scale rows."],
        ["3. Use yyyy-mm-dd for agreement dates."],
        ["4. Use numbers only for royalty scale amounts and percentages. Example: 12 for 12%."],
        ["5. Import this workbook via Admin > Data Integrity & Franchise Master."],
        ["6. The importer matches by Franchise ID first, then Franchise Code, then Business Name."],
    ]
    for row in lines:
        ws.append(row)
    ws["A1"].font = Font(size=16, bold=True, color="153D2A")
    ws.column_dimensions["A"].width = 120


def import_franchise_master_workbook(file_storage) -> Dict[str, Any]:
    wb = load_workbook(file_storage, data_only=True)
    if "Franchise Master" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'Franchise Master' sheet.")
    ws = wb["Franchise Master"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    header_map = {h: idx + 1 for idx, h in enumerate(headers)}
    missing = [h for h in MASTER_HEADERS[:19] if h not in header_map]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    lookups = franchise_lookup()
    updated = 0
    unmatched = []
    scale_updates = 0
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, header_map[h]).value for h in headers if h}
        if not any(row.get(h) for h in ("Franchise ID", "Business Name", "Franchise Code")):
            continue
        franchise = find_franchise(row, lookups)
        if not franchise:
            unmatched.append({"row": r, "business_name": row.get("Business Name"), "franchise_code": row.get("Franchise Code")})
            continue
        franchise.business_name = str(row.get("Business Name") or franchise.business_name or "").strip()
        franchise.franchise_code = str(row.get("Franchise Code") or franchise.franchise_code or "").strip()
        franchise.province = str(row.get("Province") or infer_province(franchise.business_name, row.get("Office Address")) or "Unassigned").strip()
        franchise.region = str(row.get("Region") or franchise.province or "").strip()
        franchise.district = str(row.get("District") or "").strip()
        franchise.municipality = str(row.get("Municipality") or "").strip()
        franchise.office_address = str(row.get("Office Address") or "").strip()
        franchise.office_number = str(row.get("Office Number") or "").strip()
        franchise.after_hours_number = str(row.get("After Hours Number") or "").strip()
        franchise.franchisee_name = str(row.get("Franchisee Name") or "").strip()
        franchise.franchisee_surname = str(row.get("Franchisee Surname") or "").strip()
        franchise.franchisee_cell = str(row.get("Franchisee Cell") or "").strip()
        franchise.franchisee_email = str(row.get("Franchisee Email") or "").strip().lower()
        franchise.public_email = str(row.get("Public Email") or "").strip().lower()
        franchise.agreement_start_date = parse_date(row.get("Agreement Start Date"))
        franchise.agreement_end_date = parse_date(row.get("Agreement End Date"))
        method = str(row.get("Royalty Method") or franchise.royalty_gross_method or "old").lower().strip()
        franchise.royalty_gross_method = method if method in ("old", "new") else "old"
        franchise.minimum_royalty_amount = decimal_or_zero(row.get("Minimum Royalty Amount"))
        updated += 1

        # Replace scale rows only when at least one scale row exists in the workbook.
        parsed_scales = []
        for idx in range(1, ROYALTY_SCALE_COUNT + 1):
            amount_from = decimal_or_zero(row.get(f"Scale {idx} From"))
            amount_to = decimal_or_zero(row.get(f"Scale {idx} To"))
            percent = decimal_or_zero(row.get(f"Scale {idx} %"))
            if amount_from or amount_to or percent:
                parsed_scales.append((idx, amount_from, amount_to, percent))
        if parsed_scales:
            RoyaltyScale.query.filter_by(franchise_id=franchise.id).delete()
            for idx, amount_from, amount_to, percent in parsed_scales:
                db.session.add(RoyaltyScale(franchise_id=franchise.id, row_number=idx, amount_from=amount_from, amount_to=amount_to, percentage=percent))
                scale_updates += 1
    db.session.commit()
    return {"updated": updated, "unmatched": unmatched, "scale_rows": scale_updates}
