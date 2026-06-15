from pathlib import Path
p=Path('/mnt/data/v19work/app/monthly/routes.py')
s=p.read_text()
# Ensure imports
s=s.replace('from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, abort, send_file', 'from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, abort, send_file')
# add imports if not present
if 'from werkzeug.utils import secure_filename' not in s:
    pass
# inject helpers before create_monthly_figure_from_pdf
insert = r'''
EXCEL_MONTH_NAMES = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "mei": 5,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "okt": 10,
    "oct": 10,
    "nov": 11,
    "des": 12,
    "dec": 12,
}

EXCEL_VALUE_ROWS = {
    "funeral_receipts": 6,
    "claim_receipts": 7,
    "society_receipts": 8,
    "cash_sales": 9,
    "tombstone_receipts": 10,
    "obo_service_receipts": 11,
    "insurance_receipts": 13,
    "insurance_payover": 14,
    "insurance_joinings": 16,
    "mf_files": 17,
}


def normalize_franchise_key(value):
    text = str(value or "").strip()
    text = re.sub(r"\(\s*f\s*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^a-z0-9]+", " ", text.lower())
    return " ".join(text.split())


def clean_excel_franchise_name(value):
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    if not text or text.upper() == "TOTAL":
        return ""
    text = re.sub(r"\(\s*f\s*\)", "", text, flags=re.IGNORECASE).strip()
    return text


def parse_excel_sheet_period(sheet_name):
    match = re.match(r"\s*([A-Za-zÀ-ÿ]+)\s*'?\s*(\d{2,4})\s*$", str(sheet_name or ""))
    if not match:
        return None, None
    month_key = match.group(1).strip().lower()[:3]
    month = EXCEL_MONTH_NAMES.get(month_key)
    if not month:
        return None, None
    year_value = int(match.group(2))
    year = 2000 + year_value if year_value < 100 else year_value
    return month, year


def excel_decimal(value):
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        if isinstance(value, str):
            value = value.replace("R", "").replace(",", "").replace(" ", "").strip()
        return Decimal(str(value or "0"))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def find_or_create_franchise_from_excel(raw_name):
    cleaned_name = clean_excel_franchise_name(raw_name)
    key = normalize_franchise_key(cleaned_name)
    if not key:
        return None, False

    franchises = Franchise.query.all()
    for franchise in franchises:
        if normalize_franchise_key(franchise.business_name) == key:
            return franchise, False

    franchise = Franchise(
        business_name=cleaned_name,
        franchise_code=re.sub(r"[^A-Z0-9]+", "", cleaned_name.upper())[:20],
    )
    db.session.add(franchise)
    db.session.flush()
    return franchise, True


def find_or_create_franchise_user_for_franchise(franchise):
    from app.admin.routes import get_or_create_role, slugify_email_part, temporary_password

    email = f"{slugify_email_part(franchise.business_name)}@martinsdirect.com"
    user = User.query.filter(db.func.lower(User.email) == email.lower()).first()
    created = False
    if not user:
        user = User(
            name=franchise.business_name,
            surname="User",
            email=email,
            is_active=True,
            is_active_account=True,
        )
        user.set_password(temporary_password())
        db.session.add(user)
        db.session.flush()
        created = True

    role = get_or_create_role("Franchise User")
    if role not in user.roles:
        user.roles.append(role)

    if franchise not in user.assigned_franchises:
        user.assigned_franchises.append(franchise)

    return user, created


def row_has_excel_data(values):
    return any(excel_decimal(value) != 0 for value in values.values())


def import_monthly_figures_excel_file(file_storage, allocate_users=True):
    suffix = Path(file_storage.filename or "").suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Please upload an Excel workbook (.xlsx or .xlsm).")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel import needs openpyxl. Run: pip install openpyxl") from exc

    workbook = load_workbook(file_storage, read_only=True, data_only=True)

    imported = 0
    updated = 0
    skipped = 0
    franchises_created = 0
    users_created = 0
    users_linked = 0
    periods = set()
    franchise_names = set()

    for worksheet in workbook.worksheets:
        month, year = parse_excel_sheet_period(worksheet.title)
        if not month or not year:
            skipped += 1
            continue

        periods.add(f"{year}-{month:02d}")
        max_column = worksheet.max_column or 0
        for column in range(3, max_column + 1):
            raw_name = worksheet.cell(row=2, column=column).value
            franchise_name = clean_excel_franchise_name(raw_name)
            if not franchise_name:
                continue

            values = {field: worksheet.cell(row=row_number, column=column).value for field, row_number in EXCEL_VALUE_ROWS.items()}
            if not row_has_excel_data(values):
                skipped += 1
                continue

            franchise, created_franchise = find_or_create_franchise_from_excel(franchise_name)
            if not franchise:
                skipped += 1
                continue
            if created_franchise:
                franchises_created += 1
            franchise_names.add(franchise.business_name)

            if allocate_users:
                before_links = len(franchise.assigned_users)
                user, created_user = find_or_create_franchise_user_for_franchise(franchise)
                if created_user:
                    users_created += 1
                if len(franchise.assigned_users) > before_links or franchise in user.assigned_franchises:
                    users_linked += 1

            monthly_figure = MonthlyFigure.query.filter_by(franchise_id=franchise.id, month=month, year=year).first()
            if monthly_figure:
                updated += 1
            else:
                monthly_figure = MonthlyFigure(
                    franchise_id=franchise.id,
                    month=month,
                    year=year,
                    created_by_id=current_user.id,
                    status="Imported",
                )
                db.session.add(monthly_figure)
                imported += 1

            monthly_figure.funeral_receipts = excel_decimal(values.get("funeral_receipts"))
            monthly_figure.claim_receipts = excel_decimal(values.get("claim_receipts"))
            monthly_figure.society_receipts = excel_decimal(values.get("society_receipts"))
            monthly_figure.cash_sales = excel_decimal(values.get("cash_sales"))
            monthly_figure.tombstone_receipts = excel_decimal(values.get("tombstone_receipts"))
            monthly_figure.obo_service_receipts = excel_decimal(values.get("obo_service_receipts"))
            monthly_figure.insurance_receipts = excel_decimal(values.get("insurance_receipts"))
            monthly_figure.insurance_payover = excel_decimal(values.get("insurance_payover"))
            monthly_figure.insurance_joinings = parse_int(values.get("insurance_joinings"))
            monthly_figure.mf_files = parse_int(values.get("mf_files"))
            monthly_figure.notes = f"Imported from Excel workbook: {file_storage.filename}; Sheet: {worksheet.title}; Franchise: {franchise_name}"
            if monthly_figure.status == "Draft":
                monthly_figure.status = "Imported"
            recalculate_monthly_figure(monthly_figure)

    db.session.commit()
    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "franchises_created": franchises_created,
        "users_created": users_created,
        "users_linked": users_linked,
        "period_count": len(periods),
        "franchise_count": len(franchise_names),
        "first_period": sorted(periods)[0] if periods else "",
        "last_period": sorted(periods)[-1] if periods else "",
    }

'''
if 'EXCEL_MONTH_NAMES' not in s:
    marker='def create_monthly_figure_from_pdf(file_storage, franchise_id=None):'
    s=s.replace(marker, insert + '\n' + marker)
# add route before import_pdf
route = r'''
@monthly_bp.route("/import-excel", methods=["GET", "POST"])
@login_required
@permission_required("monthly_figures:import")
def import_excel():
    if request.method == "POST":
        file_storage = request.files.get("excel_file")
        allocate_users = request.form.get("allocate_users") == "yes"
        if not file_storage or not file_storage.filename:
            flash("Please select the Excel workbook to import.", "warning")
            return redirect(url_for("monthly.import_excel"))
        try:
            result = import_monthly_figures_excel_file(file_storage, allocate_users=allocate_users)
        except Exception as exc:
            db.session.rollback()
            flash(str(exc), "danger")
            return redirect(url_for("monthly.import_excel"))

        log_action(
            "Monthly Figures",
            "Imported Excel monthly figures",
            f"New: {result['imported']}, Updated: {result['updated']}, Franchises: {result['franchise_count']}, Periods: {result['period_count']}",
        )
        db.session.commit()
        flash(
            f"Excel import complete. {result['imported']} new records created, {result['updated']} records updated, "
            f"{result['franchise_count']} franchises allocated across {result['period_count']} period(s).",
            "success",
        )
        return render_template("monthly/import_excel.html", result=result)

    return render_template("monthly/import_excel.html", result=None)

'''
if '@monthly_bp.route("/import-excel"' not in s:
    s=s.replace('@monthly_bp.route("/import", methods=["GET", "POST"])', route+'@monthly_bp.route("/import", methods=["GET", "POST"])')
p.write_text(s)
